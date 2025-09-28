
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
from .forms import *
import logging
from django.contrib.auth.decorators import login_required
from .models import Projects, Certificate, student, LeetCode, Faculty,FillOutForm, FillOutField, student, Technology, AuditLog, Achievement, AchievementCategory, ApprovalWorkflow, AcademicPerformance, EnhancedNotification, NotificationTemplate, PlacementOffer
import requests
from django.http import HttpResponse, JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
from django.core import serializers
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from django.core.mail import send_mail
import random
from flex import settings
from django.core.files.storage import FileSystemStorage
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from .serializers import StudentSerializer, ProjectSerializer, CertificateSerializer, TechnologySerializer
from django.db.models import F, Sum, Count, Avg
from .forms import PlacementOfferForm
from django.utils import timezone
from datetime import timedelta

# Faculty: Add placement offer for any student
@login_required
def add_placement_offer_faculty(request):
    if request.user.type() != "Faculty":
        return HttpResponse("Unauthorized", status=403)
    if request.method == 'POST':
        form = PlacementOfferForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Placement offer added successfully!")
            return redirect('faculty_dashboard')  # Change to your faculty dashboard name
    else:
        form = PlacementOfferForm()
    return render(request, 'add_placement_offer_faculty.html', {'form': form})

# Student: Add placement offer for self
@login_required
def add_placement_offer_student(request):
    if request.user.type() != "student":
        return HttpResponse("Unauthorized", status=403)
    if request.method == 'POST':
        form = PlacementOfferForm(request.POST)
        if form.is_valid():
            offer = form.save(commit=False)
            offer.student = request.user
            offer.save()
            messages.success(request, "Placement offer added successfully!")
            return redirect('dashboard')
    else:
        form = PlacementOfferForm()
        form.fields['student'].widget = forms.HiddenInput()
    return render(request, 'add_placement_offer_student.html', {'form': form})
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse

# Export Placement Report View
@csrf_exempt
def export_placement_report(request):
    if request.method == 'POST':
        report_type = request.POST.get('type')
        # Dummy implementation: return a simple response for now
        if report_type == 'college':
            return HttpResponse('College Summary Report (dummy)', content_type='text/plain')
        elif report_type == 'dept':
            return HttpResponse('Department Summary Report (dummy)', content_type='text/plain')
        elif report_type == 'excel':
            return HttpResponse('Excel Export (dummy)', content_type='text/plain')
        elif report_type == 'pdf':
            return HttpResponse('PDF Export (dummy)', content_type='text/plain')
        else:
            return HttpResponse('Unknown report type', content_type='text/plain')
    return HttpResponse('Invalid request', content_type='text/plain')


def index(request):
    if request.user.is_authenticated:
        print(request.user)
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        try:
            if hasattr(request.user, 'type') and request.user.type() == "student":
                return redirect('dashboard')
            elif hasattr(request.user, 'type') and request.user.type() == "Faculty":
                return redirect('faculty')
            else:
                # If user doesn't have type method, redirect to dashboard
                return redirect('dashboard')
        except:
            # Fallback to dashboard for any errors
            return redirect('dashboard')
    print("redirecting to Login")
    return redirect('login')

logging.basicConfig(level=logging.DEBUG)

# REST API endpoints
@api_view(['GET'])
@permission_classes([AllowAny])
def api_overview(request):
    api_urls = {
        'Student List': '/api/students/',
        'Student Detail': '/api/student/<str:rollno>/',
        'Projects': '/api/projects/',
        'Certificates': '/api/certificates/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([AllowAny])
def student_list(request):
    students = student.objects.values('id', 'roll_no', 'dept', 'section', 'first_name')
    serializer = StudentSerializer(students, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([AllowAny])
def student_detail(request, rollno):
    student_obj = get_object_or_404(student, rollno=rollno)
    serializer = StudentSerializer(student_obj)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
def api_login(request):
    rollno = request.data.get('username')
    password = request.data.get('password')

    if not rollno or not password:
        return Response({'error': 'Please provide both username and password'}, status=400)

    user = authenticate(request=request, username=rollno, password=password)

    if user is not None:
        login(request, user)
        try:
            if user.type() == "student":
                serializer = StudentSerializer(user)
                return Response(serializer.data)
            else:
                return Response({'error': 'Only student login supported via API'}, status=403)
        except AttributeError:
            return Response({'error': 'User type not supported'}, status=500)
    else:
        # Log the failed attempt for debugging
        logging.debug(f"Failed login attempt for username: {rollno}")
        return Response({'error': 'Invalid credentials. Please check your username and password.'}, status=401)

@api_view(['GET'])
@permission_classes([AllowAny])  # Allow any access to check auth status
def api_current_user(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Not authenticated'}, status=401)

    try:
        user_type = request.user.type()
        if user_type == "student":
            serializer = StudentSerializer(request.user)
            return Response(serializer.data)
        else:
            return Response({'error': 'Currently only student accounts are supported'}, status=403)
    except AttributeError:
        return Response({'error': 'User type not supported'}, status=403)

@api_view(['POST'])
@permission_classes([AllowAny])  # Allow any user to logout
@csrf_exempt
def api_logout(request):
    if request.user.is_authenticated:
        logout(request)
    return Response({"message": "Successfully logged out"})

@api_view(['GET'])
@permission_classes([AllowAny])
def technology_list(request):
    technologies = Technology.objects.all()
    serializer = TechnologySerializer(technologies, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([AllowAny])
def project_list(request):
    projects = Projects.objects.all()
    serializer = ProjectSerializer(projects, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_project_api(request):
    """API endpoint for creating a new project"""
    try:
        # Get current user
        current_user = request.user

        # Check if the user is a student
        if not hasattr(current_user, 'type') or current_user.type() != "student":
            return Response({"error": "Only students can create projects"}, status=403)

        # Create a form with the request data
        form = ProjectsForm(request.data, student=current_user)

        if form.is_valid():
            project = form.save(commit=False)
            project.save()

            # Save many-to-many relationships
            form.save_m2m()

            # Add current user as contributor if not already added
            if current_user not in project.contributors.all():
                project.contributors.add(current_user)

            serializer = ProjectSerializer(project)
            return Response(serializer.data, status=201)
        else:
            return Response({"errors": form.errors}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def certificate_list(request):
    certificates = Certificate.objects.all()
    serializer = CertificateSerializer(certificates, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_certificate_api(request):
    """API endpoint for creating a new certificate"""
    try:
        # Get current user
        current_user = request.user

        # Check if the user is a student
        if not hasattr(current_user, 'type') or current_user.type() != "student":
            return Response({"error": "Only students can add certificates"}, status=403)

        # Handle file upload if present
        certificate_file = request.FILES.get('certificate')

        # Create a copy of the data that can be modified
        data = request.data.copy()

        # Create a form with the request data
        form = CertificateForm(data, request.FILES, student=current_user)

        if form.is_valid():
            certificate = form.save(commit=False)
            certificate.rollno = current_user
            certificate.save()

            # Save many-to-many relationships
            form.save_m2m()

            serializer = CertificateSerializer(certificate)
            return Response(serializer.data, status=201)
        else:
            return Response({"errors": form.errors}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)

def getStudentDetails(student):
    try:
        # Calculate profile completion
        profile_completion = student.calculate_profile_completion()
        student.save()
        
        # Get academic performance
        academic_records = AcademicPerformance.objects.filter(student=student).order_by('-year', '-semester')
        latest_academic = academic_records.first() if academic_records.exists() else None
        
        # Get achievements with approval status
        achievements = Achievement.objects.filter(student=student).order_by('-submission_date')
        achievements_summary = {
            'total': achievements.count(),
            'approved': achievements.filter(status='approved').count(),
            'pending': achievements.filter(status='pending').count(),
            'rejected': achievements.filter(status='rejected').count(),
        }
        
        # Ensure we are querying the correct model field
        projects = Projects.objects.filter(contributors=student)
        
        # Get certificates by category
        technical_certificates = Certificate.objects.filter(rollno=student, category="technical")
        foreign_language_certificates = Certificate.objects.filter(rollno=student, category="foreign_language")
        co_curricular_certificates = Certificate.objects.filter(rollno=student, category="co_curricular")
        extra_curricular_certificates = Certificate.objects.filter(rollno=student, category="extra_curricular")
        
        # Debug: Log certificate counts
        logging.info(f"Certificates for {student}: Technical={technical_certificates.count()}, "
                    f"Foreign={foreign_language_certificates.count()}, "
                    f"Co-curricular={co_curricular_certificates.count()}, "
                    f"Extra-curricular={extra_curricular_certificates.count()}")
        
        # Get projects summary
        projects_summary = {
            'total': projects.count(),
            'completed': projects.filter(status='Completed').count(),
            'in_progress': projects.filter(status='In_progress').count(),
            'initialized': projects.filter(status='Initialized').count(),
        }
        
        # Get LeetCode stats
        try:
            leetcode_data = LeetCode.objects.get(rollno=student)
        except LeetCode.DoesNotExist:
            leetcode_data = LeetCode.objects.create(rollno=student)
        
        # Get placement offers
        placement_offers = PlacementOffer.objects.filter(student=student)
        
        # Calculate analytics for dashboard charts
        try:
            # Avoid Sum aggregation initially, calculate manually to prevent 'mul' error
            approved_achievements = achievements.filter(status='approved')
            total_points = 0
            for achievement in approved_achievements:
                if achievement.points_awarded:
                    total_points += achievement.points_awarded
                    
            analytics = {
                'achievement_points': total_points,
                'monthly_activity': get_monthly_activity(student),
                'skill_distribution': get_skill_distribution(student),
                'comparative_ranking': get_comparative_ranking(student),
            }
        except Exception as e:
            # Handle any aggregation errors
            logging.error(f"Error calculating analytics: {e}")
            analytics = {
                'achievement_points': 0,
                'monthly_activity': [],
                'skill_distribution': {},
                'comparative_ranking': 0,
            }
        
        # Add academic trend data
        analytics['academic_trend'] = []
        for record in academic_records[:6]:  # Last 6 semesters
            analytics['academic_trend'].append({
                'semester': f"{record.year}-{record.semester}",
                'cgpa': float(record.cgpa) if record.cgpa else 0,
                'sgpa': float(record.sgpa) if record.sgpa else 0,
                'credits': record.credits_earned,
            })

        return {
            "name": student.first_name,
            "rollno": student.roll_no,
            "dept": student.dept,
            "section": student.section,
            "leetcode_user": student.leetcode_user,
            'student': student,
            'profile_completion': profile_completion,
            'academic_records': academic_records,
            'latest_academic': latest_academic,
            'achievements': serializers.serialize('json', achievements),
            'achievements_summary': achievements_summary,
            'achievement_points': analytics.get('achievement_points', 0),  # Add achievement_points at top level
            'projects': serializers.serialize('json', projects),
            'projects_length': len(projects),
            'projects_summary': projects_summary,
            'technical_certificates': serializers.serialize('json', technical_certificates),
            'foreign_language_certificates': serializers.serialize('json', foreign_language_certificates),
            'co_curricular_certificates': serializers.serialize('json', co_curricular_certificates),
            'extra_curricular_certificates': serializers.serialize('json', extra_curricular_certificates),
            'leetcode_data': leetcode_data,
            'placement_offers': placement_offers,
            'analytics': analytics,
            'pending_approvals': ApprovalWorkflow.objects.filter(
                student=student, 
                current_status__in=['pending', 'under_review']
            ).count(),
            'recent_notifications': EnhancedNotification.objects.filter(
                recipient_student=student,
                is_read=False
            )[:5],
            # Keep old keys for backward compatibility
            'Technical': serializers.serialize('json', technical_certificates),
            'Foreign_languages': serializers.serialize('json', foreign_language_certificates)
        }
    except Exception as e:
        logging.error(f"Error in getStudentDetails: {e}")
        return {}


def get_monthly_activity(student):
    """Get monthly activity data for the last 12 months"""
    from datetime import timedelta
    
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)
    
    # Get activity counts by month
    activities = []
    
    for i in range(12):
        # Calculate proper month boundaries
        target_month = now.month - i
        target_year = now.year
        
        # Handle year rollover
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        
        # Create proper timezone-aware month boundaries
        month_start = timezone.datetime(target_year, target_month, 1, tzinfo=now.tzinfo)
        
        # Calculate next month for end boundary
        if target_month == 12:
            next_month = 1
            next_year = target_year + 1
        else:
            next_month = target_month + 1
            next_year = target_year
            
        month_end = timezone.datetime(next_year, next_month, 1, tzinfo=now.tzinfo)
        
        # Count various activities for this month
        certificates = Certificate.objects.filter(
            rollno=student,
            uploaded_at__gte=month_start,
            uploaded_at__lt=month_end
        ).count()
        
        achievements = Achievement.objects.filter(
            student=student,
            submission_date__gte=month_start,
            submission_date__lt=month_end
        ).count()
        
        activities.append({
            'month': month_start.strftime('%b %Y'),
            'certificates': certificates,
            'achievements': achievements,
            'total': certificates + achievements
        })
    
    return activities


def get_skill_distribution(student_obj):
    """Get skill distribution from certificates and projects"""
    skills = {}
    
    # Get skills from technical certificates
    tech_certs = Certificate.objects.filter(rollno=student_obj, category='technical')
    for cert in tech_certs:
        if cert.domain:
            skills[cert.domain] = skills.get(cert.domain, 0) + 1
    
    # Get technologies from projects
    projects = Projects.objects.filter(contributors=student_obj)
    for project in projects:
        for tech in project.technologies.all():
            skills[tech.name] = skills.get(tech.name, 0) + 1
    
    # Convert to list of dictionaries for frontend
    return [{'skill': k, 'count': v} for k, v in sorted(skills.items(), key=lambda x: x[1], reverse=True)[:10]]


def get_comparative_ranking(student_obj):
    """Get student's ranking compared to peers"""
    
    # Get all students in same department and year
    peers = student.objects.filter(dept=student_obj.dept, year=student_obj.year)
    
    # Calculate various metrics
    cert_count = Certificate.objects.filter(rollno=student_obj).count()
    project_count = Projects.objects.filter(contributors=student_obj).count()
    achievement_count = Achievement.objects.filter(student=student_obj, status='approved').count()
    
    return {
        'certificates': {
            'student': cert_count,
            'percentile': calculate_percentile(peers, cert_count, 'certificates')
        },
        'projects': {
            'student': project_count,
            'percentile': calculate_percentile(peers, project_count, 'projects')
        },
        'achievements': {
            'student': achievement_count,
            'percentile': calculate_percentile(peers, achievement_count, 'achievements')
        }
    }


def calculate_percentile(peers, student_value, metric_type):
    """Calculate percentile ranking for a student"""
    if not peers.exists():
        return 50
    
    # Simple percentile calculation
    better_count = 0
    total_count = peers.count()
    
    for peer in peers:
        if metric_type == 'certificates':
            peer_value = Certificate.objects.filter(rollno=peer).count()
        elif metric_type == 'projects':
            peer_value = Projects.objects.filter(contributors=peer).count()
        elif metric_type == 'achievements':
            peer_value = Achievement.objects.filter(student=peer, status='approved').count()
        else:
            peer_value = 0
        
        if student_value > peer_value:
            better_count += 1
    
    return int((better_count / total_count) * 100) if total_count > 0 else 50


@login_required
def dashboard(request):
    try:
        context = getStudentDetails(request.user)
        
        # Add achievement categories for the modal form
        from .models import AchievementCategory
        context['achievement_categories'] = AchievementCategory.objects.filter(is_active=True)
        
        response = render(request, 'dashboard.html', context)
        # Add cache control headers to prevent caching issues
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    except Exception as e:
        logging.error(f"Error in dashboard: {e}")
        return HttpResponse("An error occurred.")


# Achievement Management Views
@login_required
def add_achievement(request):
    """Add new achievement for approval"""
    if request.method == 'POST':
        try:
            from datetime import datetime
            from django.contrib import messages
            
            # Create achievement
            achievement = Achievement.objects.create(
                student=request.user,
                submitted_by=request.user,
                category_id=request.POST.get('category'),
                title=request.POST.get('title'),
                description=request.POST.get('description'),
                achievement_date=request.POST.get('achievement_date'),
                verification_method=request.POST.get('verification_method', ''),
                status='pending'  # Set initial status as pending
            )
            
            # Handle optional fields
            if request.POST.get('additional_notes'):
                achievement.faculty_comments = request.POST.get('additional_notes')
            
            # Handle file upload
            if 'supporting_documents' in request.FILES:
                achievement.supporting_documents = request.FILES['supporting_documents']
            
            achievement.save()
            
            messages.success(request, f'Achievement "{achievement.title}" has been submitted for approval!')
            
            # Redirect to dashboard with cache busting parameter
            from django.urls import reverse
            import time
            return redirect(f"{reverse('dashboard')}?t={int(time.time())}")
            
        except Exception as e:
            logging.error(f"Error adding achievement: {e}")
            messages.error(request, f'Error adding achievement: {str(e)}')
            return redirect('dashboard')
    
    # If GET request, redirect to dashboard
    return redirect('dashboard')

@login_required
def edit_achievement(request, achievement_id):
    """Edit achievement - only allowed for pending achievements"""
    try:
        achievement = get_object_or_404(Achievement, id=achievement_id, student=request.user)
        
        # Only allow editing of pending achievements
        if achievement.status != 'pending':
            messages.error(request, 'You can only edit pending achievements.')
            return redirect('dashboard')
        
        if request.method == 'POST':
            # Update achievement fields
            achievement.title = request.POST.get('title', achievement.title)
            achievement.description = request.POST.get('description', achievement.description)
            achievement.achievement_date = request.POST.get('achievement_date', achievement.achievement_date)
            achievement.verification_method = request.POST.get('verification_method', achievement.verification_method)
            
            # Update category if provided
            if request.POST.get('category'):
                achievement.category_id = request.POST.get('category')
            
            # Handle file upload
            if 'supporting_documents' in request.FILES:
                achievement.supporting_documents = request.FILES['supporting_documents']
            
            # Handle additional notes
            if request.POST.get('additional_notes'):
                achievement.faculty_comments = request.POST.get('additional_notes')
            
            achievement.save()
            
            messages.success(request, f'Achievement "{achievement.title}" has been updated successfully!')
            return redirect('dashboard')
        
        # For GET request, render edit form
        from .models import AchievementCategory
        achievement_categories = AchievementCategory.objects.filter(is_active=True)
        
        return render(request, 'edit_achievement.html', {
            'achievement': achievement,
            'achievement_categories': achievement_categories
        })
        
    except Exception as e:
        logging.error(f"Error editing achievement: {e}")
        messages.error(request, f'Error editing achievement: {str(e)}')
        return redirect('dashboard')

@login_required
def delete_achievement(request, achievement_id):
    """Delete achievement - only allowed for pending achievements"""
    try:
        achievement = get_object_or_404(Achievement, id=achievement_id, student=request.user)
        
        # Only allow deletion of pending achievements
        if achievement.status != 'pending':
            return JsonResponse({
                'success': False, 
                'error': 'You can only delete pending achievements.'
            })
        
        if request.method == 'POST':
            title = achievement.title
            achievement.delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Achievement "{title}" has been deleted successfully!'
            })
        
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
        
    except Exception as e:
        logging.error(f"Error deleting achievement: {e}")
        return JsonResponse({
            'success': False, 
            'error': f'Error deleting achievement: {str(e)}'
        })

@login_required
def achievement_list(request):
    """List all achievements for current student"""
    achievements = Achievement.objects.filter(student=request.user).order_by('-submission_date')
    return render(request, 'achievement_list.html', {'achievements': achievements})


# Faculty Views for Approval System
@login_required 
def faculty_approval_dashboard(request):
    """Faculty dashboard for managing approvals"""
    print("Faculty approval dashboard accessed")  # Debug log
    
    try:
        # Start with minimal data
        context = {
            'achievements': [],
            'projects': [],
            'certificates': [],
            'categories': [],
            'pending_count': 0,
            'approved_today': 0,
            'total_reviewed': 0,
            'avg_review_time': '2.5'
        }
        
        # Try to load achievements
        try:
            achievements = Achievement.objects.all().order_by('-id')[:20]  # Limit for performance
            context['achievements'] = achievements
            print(f"Loaded {len(achievements)} achievements")
        except Exception as e:
            print(f"Error loading achievements: {e}")
        
        # Try to load projects
        try:
            projects = Projects.objects.filter(approval_status='pending').order_by('-id')[:20]
            context['projects'] = projects
            print(f"Loaded {len(projects)} projects")
        except Exception as e:
            print(f"Error loading projects: {e}")
        
        # Try to load certificates
        try:
            certificates = Certificate.objects.filter(approval_status='pending').order_by('-id')[:20]
            context['certificates'] = certificates
            print(f"Loaded {len(certificates)} certificates")
        except Exception as e:
            print(f"Error loading certificates: {e}")
        
        # Try to load categories
        try:
            categories = AchievementCategory.objects.all()
            context['categories'] = categories
            print(f"Loaded {len(categories)} categories")
        except Exception as e:
            print(f"Error loading categories: {e}")
            
        # Try to get pending count
        try:
            pending_count = Achievement.objects.filter(status='pending').count()
            context['pending_count'] = pending_count
            print(f"Pending count: {pending_count}")
        except Exception as e:
            print(f"Error getting pending count: {e}")
            
        # Try to get reviewed count
        try:
            total_reviewed = Achievement.objects.exclude(status='pending').count()
            context['total_reviewed'] = total_reviewed
            print(f"Total reviewed: {total_reviewed}")
        except Exception as e:
            print(f"Error getting reviewed count: {e}")
        
        print("Rendering template...")
        return render(request, 'faculty_approval_dashboard.html', context)
        
    except Exception as e:
        import traceback
        error_msg = f'Critical error in faculty_approval_dashboard: {str(e)}\n{traceback.format_exc()}'
        print(error_msg)
        
        # Return error page with debugging info
        from django.http import HttpResponse
        return HttpResponse(f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Approval Dashboard Error</title>
            <style>
                body {{ font-family: Arial; padding: 20px; background: #1a1a1a; color: #f5f5f5; }}
                .error {{ background: #000; padding: 15px; border: 1px solid #f1c40f; border-radius: 5px; }}
                .back-btn {{ color: #f1c40f; text-decoration: none; padding: 10px 20px; border: 1px solid #f1c40f; border-radius: 5px; display: inline-block; margin-top: 20px; }}
            </style>
        </head>
        <body>
            <h1 style="color: #f1c40f;">🚫 Approval Dashboard Error</h1>
            <p>There was an error loading the approval dashboard. Debug information:</p>
            <div class="error">
                <pre>{error_msg}</pre>
            </div>
            <a href="/faculty" class="back-btn">← Back to Faculty Dashboard</a>
            
            <h3 style="color: #f1c40f; margin-top: 30px;">Troubleshooting:</h3>
            <ul>
                <li>Check if Achievement and AchievementCategory models exist</li>
                <li>Check database connections</li>
                <li>Check template file exists at: flexapp/templates/faculty_approval_dashboard.html</li>
                <li>Check console logs for JavaScript errors</li>
            </ul>
        </body>
        </html>
        """, status=500)


@login_required
def approve_achievement(request, achievement_id):
    """Approve or reject an achievement"""
    try:
        achievement = get_object_or_404(Achievement, id=achievement_id)
        
        if request.method == 'POST':
            status = request.POST.get('status')
            comments = request.POST.get('faculty_comments', '')
            points_awarded = request.POST.get('points_awarded', 0)
            
            if status == 'approved':
                achievement.status = 'approved'
                achievement.approved_by = request.user
                achievement.approval_date = timezone.now()
                achievement.faculty_comments = comments
                achievement.points_awarded = int(points_awarded) if points_awarded else 0
                achievement.is_verified = True
                
                messages.success(request, f'Achievement "{achievement.title}" approved successfully!')
                
            elif status == 'rejected':
                achievement.status = 'rejected'
                achievement.reviewed_by = request.user
                achievement.review_date = timezone.now()
                achievement.rejection_reason = comments
                
                messages.warning(request, f'Achievement "{achievement.title}" rejected.')
            
            achievement.save()
            
            return redirect('faculty_approval_dashboard')
    
    except Exception as e:
        logging.error(f"Error in approval: {e}")
        messages.error(request, 'Error processing approval. Please try again.')
    
    return redirect('faculty_approval_dashboard')


@login_required
def approve_project(request, project_id):
    """Faculty view to approve/reject projects"""
    if request.user.type() != "Faculty" and not request.user.is_superuser:
        return HttpResponse("Unauthorized", status=403)
    
    try:
        project = Projects.objects.get(id=project_id)
        
        if request.method == 'POST':
            status = request.POST.get('status')
            comments = request.POST.get('faculty_comments', '')
            
            if status == 'approved':
                project.approval_status = 'approved'
                project.approved_by = request.user
                project.approval_date = timezone.now()
                project.faculty_comments = comments
                
                messages.success(request, f'Project "{project.title}" approved successfully!')
                
            elif status == 'rejected':
                project.approval_status = 'rejected'
                project.approved_by = request.user
                project.approval_date = timezone.now()
                project.faculty_comments = comments
                
                messages.warning(request, f'Project "{project.title}" rejected.')
            
            project.save()
            return redirect('faculty_approval_dashboard')
    
    except Exception as e:
        logging.error(f"Error in project approval: {e}")
        messages.error(request, 'Error processing project approval. Please try again.')
    
    return redirect('faculty_approval_dashboard')


@login_required
def approve_certificate(request, certificate_id):
    """Faculty view to approve/reject certificates"""
    if request.user.type() != "Faculty" and not request.user.is_superuser:
        return HttpResponse("Unauthorized", status=403)
    
    try:
        certificate = Certificate.objects.get(id=certificate_id)
        
        if request.method == 'POST':
            status = request.POST.get('status')
            comments = request.POST.get('faculty_comments', '')
            
            if status == 'approved':
                certificate.approval_status = 'approved'
                certificate.approved_by = request.user
                certificate.approval_date = timezone.now()
                certificate.faculty_comments = comments
                
                messages.success(request, f'Certificate "{certificate.title}" approved successfully!')
                
            elif status == 'rejected':
                certificate.approval_status = 'rejected'
                certificate.approved_by = request.user
                certificate.approval_date = timezone.now()
                certificate.faculty_comments = comments
                
                messages.warning(request, f'Certificate "{certificate.title}" rejected.')
            
            certificate.save()
            return redirect('faculty_approval_dashboard')
    
    except Exception as e:
        logging.error(f"Error in certificate approval: {e}")
        messages.error(request, 'Error processing certificate approval. Please try again.')
    
    return redirect('faculty_approval_dashboard')


# Bulk Operations and Data Export
@login_required
def export_student_data(request):
    """Export student data in various formats"""
    if request.user.type() not in ["Faculty", "Admin"] and not request.user.is_superuser:
        return HttpResponse("Unauthorized", status=403)
    
    export_format = request.GET.get('format', 'csv')
    data_type = request.GET.get('type', 'students')
    
    try:
        if data_type == 'students':
            data = export_students_data(export_format)
        elif data_type == 'achievements':
            data = export_achievements_data(export_format)
        elif data_type == 'certificates':
            data = export_certificates_data(export_format)
        elif data_type == 'projects':
            data = export_projects_data(export_format)
        else:
            return HttpResponse("Invalid data type", status=400)
        
        return data
        
    except Exception as e:
        logging.error(f"Error exporting data: {e}")
        return HttpResponse("Error exporting data", status=500)


def export_students_data(request, format_type):
    """Export students data"""
    students = student.objects.all().values(
        'username', 'first_name', 'last_name', 'email', 'dept', 'year', 'section',
        'current_cgpa', 'total_credits', 'profile_completion_percentage'
    )
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students.csv"'
        
        df = pd.DataFrame(list(students))
        df.to_csv(response, index=False)
        return response
    
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
        
        df = pd.DataFrame(list(students))
        df.to_excel(response, index=False)
        return response
    
    elif format_type == 'json':
        return JsonResponse(list(students), safe=False)


def export_achievements_data(request, format_type):
    """Export achievements data"""
    achievements = Achievement.objects.select_related('student', 'category').values(
        'student__username', 'student__first_name', 'category__name',
        'title', 'description', 'status', 'points_awarded', 'submission_date'
    )
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="achievements.csv"'
        
        df = pd.DataFrame(list(achievements))
        df.to_csv(response, index=False)
        return response
    
    return JsonResponse(list(achievements), safe=False)



def export_certificates_data(request, format_type):
    """Export certificates data"""
    certificates = Certificate.objects.select_related('rollno').values(
        'rollno__username', 'rollno__first_name', 'title', 'category', 
        'source', 'domain', 'year_and_sem', 'uploaded_at'
    )
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="certificates.csv"'
        
        df = pd.DataFrame(list(certificates))
        df.to_csv(response, index=False)
        return response
    
    return JsonResponse(list(certificates), safe=False)


def export_projects_data(request, format_type):
    """Export projects data"""
    projects = Projects.objects.select_related().values(
        'title', 'description', 'status', 'year_and_sem', 'github_link'
    )
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="projects.csv"'
        
        df = pd.DataFrame(list(projects))
        df.to_csv(response, index=False)
        return response
    
    return JsonResponse(list(projects), safe=False)


# Portfolio Generation System
@login_required
def generate_portfolio(request):
    """Generate student portfolio in PDF/Web format"""
    try:
        student_data = getStudentDetails(request.user)
        
        # Calculate portfolio metrics
        portfolio_data = {
            'student': request.user,
            'academic_summary': get_academic_summary(request.user),
            'achievement_summary': get_achievement_summary(request.user),
            'project_highlights': get_project_highlights(request.user),
            'skill_matrix': get_skill_matrix(request.user),
            'certifications': get_certification_summary(request.user),
            'generated_date': timezone.now(),
            'verification_code': generate_verification_code(request.user)
        }
        
        format_type = request.GET.get('format', 'web')
        
        if format_type == 'pdf':
            return generate_pdf_portfolio(portfolio_data)
        else:
            return render(request, 'portfolio.html', portfolio_data)
            
    except Exception as e:
        logging.error(f"Error generating portfolio: {e}")
        messages.error(request, 'Error generating portfolio. Please try again.')
        return redirect('dashboard')


def get_academic_summary(student_obj):
    """Get academic performance summary"""
    latest_record = AcademicPerformance.objects.filter(student=student_obj).order_by('-year', '-semester').first()
    
    return {
        'current_cgpa': latest_record.cgpa if latest_record else student_obj.current_cgpa,
        'total_credits': latest_record.total_credits if latest_record else student_obj.total_credits,
        'current_year': student_obj.year,
        'department': student_obj.dept,
        'class_rank': latest_record.class_rank if latest_record else None,
        'attendance': latest_record.attendance_percentage if latest_record else None
    }


def get_achievement_summary(student_obj):
    """Get achievements summary"""
    achievements = Achievement.objects.filter(student=student_obj, status='approved')
    
    # Calculate total points manually to avoid aggregation issues
    total_points = 0
    for achievement in achievements:
        if achievement.points_awarded:
            total_points += achievement.points_awarded
    
    summary = {
        'total_achievements': achievements.count(),
        'total_points': total_points,
        'categories': {},
        'recent_achievements': achievements.order_by('-achievement_date')[:5]
    }
    
    # Group by category
    for achievement in achievements:
        category = achievement.category.name
        if category not in summary['categories']:
            summary['categories'][category] = 0
        summary['categories'][category] += 1
    
    return summary


def get_project_highlights(student_obj):
    """Get project highlights"""
    projects = Projects.objects.filter(contributors=student_obj)
    
    return {
        'total_projects': projects.count(),
        'completed_projects': projects.filter(status='Completed').count(),
        'technologies_used': list(set([tech.name for project in projects for tech in project.technologies.all()])),
        'featured_projects': projects.filter(status='Completed').order_by('-id')[:3]
    }


def get_skill_matrix(student_obj):
    """Get comprehensive skill matrix"""
    skills = {
        'technical': [],
        'programming': [],
        'certifications': [],
        'soft_skills': []
    }
    
    # Technical skills from projects
    projects = Projects.objects.filter(contributors=student_obj)
    for project in projects:
        for tech in project.technologies.all():
            if tech.name not in skills['programming']:
                skills['programming'].append(tech.name)
    
    # Skills from certificates
    certs = Certificate.objects.filter(rollno=student_obj, category='technical')
    for cert in certs:
        if cert.domain and cert.domain not in skills['technical']:
            skills['technical'].append(cert.domain)
    
    # Skills from student profile
    if student_obj.technical_skills:
        skills['technical'].extend([skill.strip() for skill in student_obj.technical_skills.split(',')])
    
    if student_obj.soft_skills:
        skills['soft_skills'] = [skill.strip() for skill in student_obj.soft_skills.split(',')]
    
    return skills


def get_certification_summary(student_obj):
    """Get certification summary"""
    certs = Certificate.objects.filter(rollno=student_obj)
    
    summary = {
        'total_certificates': certs.count(),
        'by_category': {},
        'recent_certifications': certs.order_by('-uploaded_at')[:5],
        'top_providers': {}
    }
    
    # Group by category
    for cert in certs:
        if cert.category not in summary['by_category']:
            summary['by_category'][cert.category] = 0
        summary['by_category'][cert.category] += 1
        
        if cert.source not in summary['top_providers']:
            summary['top_providers'][cert.source] = 0
        summary['top_providers'][cert.source] += 1
    
    return summary


def generate_verification_code(student_obj):
    """Generate unique verification code for portfolio"""
    import hashlib
    
    data = f"{student_obj.username}_{timezone.now().strftime('%Y%m%d')}"
    return hashlib.md5(data.encode()).hexdigest()[:8]


def generate_pdf_portfolio(portfolio_data):
    """Generate PDF portfolio using reportlab"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2E86AB')
        )
        story.append(Paragraph(f"Academic Portfolio - {portfolio_data['student'].first_name} {portfolio_data['student'].last_name}", title_style))
        story.append(Spacer(1, 12))
        
        # Student Information
        student_info = [
            ['Name:', f"{portfolio_data['student'].first_name} {portfolio_data['student'].last_name}"],
            ['Roll Number:', portfolio_data['student'].roll_no],
            ['Department:', portfolio_data['student'].dept],
            ['Year:', str(portfolio_data['student'].year)],
            ['Email:', portfolio_data['student'].email],
            ['Current CGPA:', str(portfolio_data['academic_summary']['current_cgpa'] or 'N/A')],
        ]
        
        student_table = Table(student_info, colWidths=[2*inch, 3*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(student_table)
        story.append(Spacer(1, 20))
        
        # Achievement Summary
        story.append(Paragraph("Achievement Summary", styles['Heading2']))
        achievement_data = portfolio_data['achievement_summary']
        story.append(Paragraph(f"Total Achievements: {achievement_data['total_achievements']}", styles['Normal']))
        story.append(Paragraph(f"Total Points Earned: {achievement_data['total_points']}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Project Summary
        story.append(Paragraph("Project Portfolio", styles['Heading2']))
        project_data = portfolio_data['project_highlights']
        story.append(Paragraph(f"Total Projects: {project_data['total_projects']}", styles['Normal']))
        story.append(Paragraph(f"Completed Projects: {project_data['completed_projects']}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Verification
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Verification Code: {portfolio_data['verification_code']}", styles['Normal']))
        story.append(Paragraph(f"Generated on: {portfolio_data['generated_date'].strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="portfolio_{portfolio_data["student"].username}.pdf"'
        return response
        
    except ImportError:
        messages.error(request, 'PDF generation not available. Please install reportlab.')
        return redirect('dashboard')
    except Exception as e:
        logging.error(f"Error generating PDF: {e}")
        messages.error(request, 'Error generating PDF portfolio.')
        return redirect('dashboard')


# Compliance and Accreditation Reports
@login_required
def compliance_reports(request):
    """Generate compliance reports for NAAC, NBA, AICTE, NIRF"""
    if not request.user.is_superuser and request.user.type() != "Faculty":
        return HttpResponse("Unauthorized", status=403)
    
    report_type = request.GET.get('type', 'naac')
    year = request.GET.get('year', timezone.now().year)
    
    try:
        if report_type == 'naac':
            data = generate_naac_report(year)
        elif report_type == 'nba':
            data = generate_nba_report(year)
        elif report_type == 'aicte':
            data = generate_aicte_report(year)
        elif report_type == 'nirf':
            data = generate_nirf_report(year)
        else:
            data = generate_custom_report(request)
        
        return render(request, 'compliance_reports.html', {
            'report_type': report_type,
            'year': year,
            'data': data,
            'generated_at': timezone.now()
        })
        
    except Exception as e:
        logging.error(f"Error generating compliance report: {e}")
        messages.error(request, 'Error generating compliance report.')
        return redirect('admin_dashboard')


def generate_naac_report(year):
    """Generate NAAC compliance report"""
    total_students = student.objects.filter(admission_year=year).count()
    
    # Student participation metrics
    participation_metrics = {
        'total_students': total_students,
        'students_with_achievements': Achievement.objects.filter(
            student__admission_year=year, 
            status='approved'
        ).values('student').distinct().count(),
        'students_with_certifications': Certificate.objects.filter(
            rollno__admission_year=year
        ).values('rollno').distinct().count(),
        'students_with_projects': Projects.objects.filter(
            contributors__admission_year=year
        ).values('contributors').distinct().count(),
    }
    
    # Calculate participation percentages
    participation_metrics['achievement_participation_rate'] = (
        participation_metrics['students_with_achievements'] / total_students * 100
    ) if total_students > 0 else 0
    
    participation_metrics['certification_participation_rate'] = (
        participation_metrics['students_with_certifications'] / total_students * 100
    ) if total_students > 0 else 0
    
    # Quality metrics
    quality_metrics = {
        'avg_cgpa': student.objects.filter(admission_year=year).aggregate(
            Avg('current_cgpa')
        )['current_cgpa__avg'] or 0,
        'students_above_8_cgpa': student.objects.filter(
            admission_year=year, 
            current_cgpa__gte=8.0
        ).count(),
        'placement_rate': calculate_placement_rate(year),
    }
    
    # Co-curricular activities
    cocurricular_metrics = {
        'technical_events': Achievement.objects.filter(
            student__admission_year=year,
            category__name__icontains='technical'
        ).count(),
        'cultural_events': Achievement.objects.filter(
            student__admission_year=year,
            category__name__icontains='cultural'
        ).count(),
        'sports_events': Achievement.objects.filter(
            student__admission_year=year,
            category__name__icontains='sports'
        ).count(),
    }
    
    return {
        'participation_metrics': participation_metrics,
        'quality_metrics': quality_metrics,
        'cocurricular_metrics': cocurricular_metrics,
        'year': year
    }


def generate_nba_report(year):
    """Generate NBA outcome-based education report"""
    students = student.objects.filter(admission_year=year)
    
    # Program Outcomes assessment
    program_outcomes = {
        'engineering_knowledge': assess_engineering_knowledge(students),
        'problem_analysis': assess_problem_analysis(students),
        'design_solutions': assess_design_solutions(students),
        'research_skills': assess_research_skills(students),
        'modern_tools': assess_modern_tools(students),
        'professional_ethics': assess_professional_ethics(students),
        'communication': assess_communication_skills(students),
        'project_management': assess_project_management(students),
        'lifelong_learning': assess_lifelong_learning(students),
    }
    
    # Course outcomes mapping
    course_outcomes = map_course_outcomes(students)
    
    # Assessment methods
    assessment_data = {
        'continuous_assessment': get_continuous_assessment_data(students),
        'project_assessment': get_project_assessment_data(students),
        'industry_assessment': get_industry_assessment_data(students),
    }
    
    return {
        'program_outcomes': program_outcomes,
        'course_outcomes': course_outcomes,
        'assessment_data': assessment_data,
        'year': year
    }


def calculate_placement_rate(year):
    """Calculate placement rate for given year"""
    total_students = student.objects.filter(
        admission_year=year,
        graduation_status='current'
    ).count()
    
    placed_students = PlacementOffer.objects.filter(
        student__admission_year=year
    ).values('student').distinct().count()
    
    return (placed_students / total_students * 100) if total_students > 0 else 0


# Helper functions for NBA assessment
def assess_engineering_knowledge(students):
    """Assess engineering knowledge through technical certifications and projects"""
    total = students.count()
    if total == 0:
        return 0
    
    qualified = 0
    for student in students:
        tech_certs = Certificate.objects.filter(rollno=student, category='technical').count()
        projects = Projects.objects.filter(contributors=student).count()
        
        # Criteria: At least 2 technical certifications OR 3 completed projects
        if tech_certs >= 2 or projects >= 3:
            qualified += 1
    
    return (qualified / total) * 100


def assess_problem_analysis(students):
    """Assess problem analysis through project complexity and research"""
    total = students.count()
    if total == 0:
        return 0
    
    qualified = 0
    for student in students:
        complex_projects = Projects.objects.filter(
            contributors=student,
            status='Completed'
        ).count()
        
        research_achievements = Achievement.objects.filter(
            student=student,
            category__name__icontains='research',
            status='approved'
        ).count()
        
        if complex_projects >= 2 or research_achievements >= 1:
            qualified += 1
    
    return (qualified / total) * 100


# Continue with other assessment functions...
def assess_design_solutions(students):
    return 75  # Placeholder

def assess_research_skills(students):
    return 70  # Placeholder

def assess_modern_tools(students):
    return 80  # Placeholder

def assess_professional_ethics(students):
    return 85  # Placeholder

def assess_communication_skills(students):
    return 75  # Placeholder

def assess_project_management(students):
    return 70  # Placeholder

def assess_lifelong_learning(students):
    return 90  # Placeholder

def map_course_outcomes(students):
    return {}  # Placeholder

def get_continuous_assessment_data(students):
    return {}  # Placeholder

def get_project_assessment_data(students):
    return {}  # Placeholder

def get_industry_assessment_data(students):
    return {}  # Placeholder

def generate_aicte_report(year):
    """Generate AICTE compliance report"""
    return {'message': 'AICTE report coming soon'}

def generate_nirf_report(year):
    """Generate NIRF ranking report"""
    return {'message': 'NIRF report coming soon'}

def generate_custom_report(request):
    """Generate custom report based on user requirements"""
    return {'message': 'Custom report builder coming soon'}

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
import logging
from .models import Projects, Technology, student

@login_required
def search_students(request):
    if "term" in request.GET:
        query = request.GET.get("term", "")
        print(query)
        students = student.objects.filter(first_name__icontains=query)[:10]  # Limit results
        students_list = [{"id": s.id, "text": s.first_name} for s in students]
        return JsonResponse(students_list, safe=False)
    return JsonResponse([], safe=False)

def search_technologies(request):
    term = request.GET.get("term", "")
    technologies = Technology.objects.filter(name__icontains=term).values("id", "name")
    return JsonResponse(list(technologies), safe=False)

@csrf_exempt
def create_technology(request):
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            
            if not name:
                return JsonResponse({'error': 'Technology name is required'}, status=400)
            
            # Check if technology already exists (case-insensitive)
            existing_tech = Technology.objects.filter(name__iexact=name).first()
            if existing_tech:
                return JsonResponse({
                    'id': existing_tech.id,
                    'name': existing_tech.name,
                    'message': 'Technology already exists'
                })
            
            # Create new technology
            new_tech = Technology.objects.create(name=name.capitalize())
            return JsonResponse({
                'id': new_tech.id,
                'name': new_tech.name,
                'message': 'Technology created successfully'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Only POST method allowed'}, status=405)

def clean_list(lst_str):
    return list(map(int, ast.literal_eval(lst_str)[0].split(',')))

import ast
import logging
import requests
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Projects, student, Technology

def extract_languages_from_github(github_url, project_instance):
    try:
        if "github.com/" not in github_url:
            return

        parts = github_url.strip('/').split("github.com/")[-1].split('/')
        if len(parts) < 2:
            return

        owner, repo = parts[0], parts[1]
        api_url = f"https://api.github.com/repos/{owner}/{repo}/languages"
        response = requests.get(api_url)
        if response.status_code != 200:
            return

        data = response.json()
        languages = list(data.keys())
        print(languages)

        for lang in languages:
            tech_obj, _ = Technology.objects.get_or_create(name=lang.capitalize())
            project_instance.technologies.add(tech_obj)

    except Exception as e:
        print(f"Error extracting languages: {e}")


@login_required
def create_project(request):
    try:
        if request.method == 'POST':
            title = request.POST.get('project-name')
            description = request.POST.get('description')
            year_and_sem = request.POST.get('year-and-sem')
            status = request.POST.get('status', 'Initialized')
            github_link = request.POST.get('github')
            contributors_ids = request.POST.getlist('contributors')
            print(type(contributors_ids))
            tech_names = request.POST.getlist('technologies')

            from itertools import chain
            contributors_ids = list(
                chain.from_iterable(id_str.split(',') for id_str in contributors_ids)
            )
            contributors_ids = [int(id.strip()) for id in contributors_ids if id.strip().isdigit()]
            print(contributors_ids)

            tech_names = [tech.strip().capitalize() for tech in tech_names if tech.strip()]

            # Step 1: Create project first
            new_project = Projects.objects.create(
                title=title,
                description=description,
                year_and_sem=year_and_sem,
                github_link=github_link,
                status=status
            )

            # Step 2: Add current user if student
            if request.user.type() == "student":
                new_project.contributors.add(request.user.id)

            # Step 3: Add additional contributors
            contributors = student.objects.filter(id__in=contributors_ids)
            new_project.contributors.add(*contributors)

            # Step 4: Add selected technologies
            tech_objects = []
            for tech_name in tech_names:
                tech_obj, _ = Technology.objects.get_or_create(name=tech_name)
                tech_objects.append(tech_obj)
            new_project.technologies.add(*tech_objects)

            # Step 5: Extract GitHub languages if none selected
            if not tech_objects and github_link:
                extract_languages_from_github(github_link, new_project)

            new_project.save()
            print(new_project," saved")
            return redirect('dashboard')

        # GET request
        students = student.objects.all()
        technologies = Technology.objects.all()
        return render(request, 'dashboard.html', {'students': students, 'technologies': technologies})

    except Exception as e:
        logging.error(f"Error in create_project: {e}")
        return HttpResponse("An error occurred.")

@login_required
def add_certification(request):
    try:
        if request.method == 'POST':
            rollno = request.user
            
            # Get form data with validation
            category = request.POST.get('category', 'technical')
            title = request.POST.get('title')
            source = request.POST.get('source')
            year_and_sem = request.POST.get('year_and_sem')
            course_link = request.POST.get('course_link', '')
            domain = request.POST.get('domain', '')
            certificate_id = request.POST.get('certificate_id', '')
            validity_period = request.POST.get('validity_period', '')
            
            # Category-specific fields
            technologies = request.POST.get('technologies', '')  # For technical certificates
            language_level = request.POST.get('language_level', '')  # For language certificates
            event_type = request.POST.get('event_type', '')  # For co-curricular/extra-curricular
            fest_name = request.POST.get('fest_name', '')
            recognition = request.POST.get('recognition', '')
            rank = request.POST.get('rank', '')
            
            # Validate required fields
            if not title or not title.strip():
                messages.error(request, "Certificate title is required.")
                return redirect('dashboard')
            
            if not source or not source.strip():
                messages.error(request, "Source/Institution is required.")
                return redirect('dashboard')
                
            if not year_and_sem:
                messages.error(request, "Year and Semester is required.")
                return redirect('dashboard')
                
            if not category:
                messages.error(request, "Category is required.")
                return redirect('dashboard')
            
            # Log the form data for debugging
            logging.info(f"Creating certificate with data: title={title}, source={source}, category={category}, year_and_sem={year_and_sem}")
            
            # Convert rank to integer if provided
            rank_int = None
            if rank and rank.strip():
                try:
                    rank_int = int(rank)
                except ValueError:
                    messages.error(request, "Rank must be a valid number.")
                    return redirect('dashboard')
            
            # Create certificate instance with only required fields first
            print(f"DEBUG: About to create certificate with fields: rollno={rollno}, title={title}, source={source[:50]}, category={category}, year_and_sem={year_and_sem}")
            new_certificate = Certificate(
                rollno=rollno,
                title=title,
                source=source[:50],  # Truncate to max_length=50
                category=category,
                year_and_sem=year_and_sem,
            )
            
            # Set optional fields separately (only fields that exist in the model)
            if course_link:
                new_certificate.course_link = course_link
            if domain:
                new_certificate.domain = domain
            if certificate_id:
                new_certificate.certificate_id = certificate_id
            if validity_period:
                new_certificate.validity_period = validity_period
            
            # Handle category-specific data
            if category == 'technical' and technologies:
                # For technical certificates, we can store technologies info in domain field
                if domain:
                    new_certificate.domain = f"{domain} - Technologies: {technologies}"
                else:
                    new_certificate.domain = f"Technologies: {technologies}"
            elif category == 'foreign_language' and language_level:
                # For language certificates, store level info in domain
                if domain:
                    new_certificate.domain = f"{domain} - Level: {language_level}"
                else:
                    new_certificate.domain = f"Language Level: {language_level}"
            elif category in ['co_curricular', 'extra_curricular']:
                # For event-based certificates, compile event info
                event_info = []
                if event_type:
                    event_info.append(f"Type: {event_type}")
                if fest_name:
                    event_info.append(f"Event: {fest_name}")
                if recognition:
                    event_info.append(f"Recognition: {recognition}")
                if rank_int:
                    event_info.append(f"Rank: {rank_int}")
                
                if event_info:
                    event_details = " | ".join(event_info)
                    if domain:
                        new_certificate.domain = f"{domain} - {event_details}"
                    else:
                        new_certificate.domain = event_details
            
            # Handle file upload
            if 'certificate' in request.FILES:
                new_certificate.certificate = request.FILES['certificate']
            
            # Save the certificate to database
            try:
                new_certificate.save()
                # Debug: Log the certificate creation
                logging.info(f"Certificate created successfully: {new_certificate.title} for user {rollno} with ID: {new_certificate.id}")
            except Exception as save_error:
                logging.error(f"Error saving certificate to database: {save_error}")
                messages.error(request, f"Failed to save certificate: {str(save_error)}")
                return redirect('dashboard')
            
            messages.success(request, f"{category.replace('_', ' ').title()} certificate added successfully!")
            # Add a timestamp to prevent caching issues
            from django.urls import reverse
            from django.http import HttpResponseRedirect
            import time
            dashboard_url = reverse('dashboard')
            return HttpResponseRedirect(f"{dashboard_url}?t={int(time.time())}")
            
        return render(request, 'dashboard.html')
    except Exception as e:
        logging.error(f"Error in add_certification: {e}")
        messages.error(request, f"An error occurred while adding the certificate: {str(e)}")
        return redirect('dashboard')

def CustomLogin(request):
    print("LoginBlock")
    try:
        if request.user.is_authenticated:
            print(request.user)
            if request.user.is_superuser:
                return redirect('admin_dashboard')
            try:
                if hasattr(request.user, 'type') and request.user.type() == "student":
                    return redirect('dashboard')
                elif hasattr(request.user, 'type') and request.user.type() == "Faculty":
                    return redirect('faculty')
            except:
                # If user doesn't have type method, redirect to dashboard
                return redirect('dashboard')

        if request.method == 'POST':
            rollno = request.POST.get('rollno')
            password = request.POST.get('password')
            print(rollno,password)
            user = authenticate(request=request, username=rollno, password=password)
            print(user)
            if user is not None:
                login(request, user)
                try:
                    if hasattr(user, 'type') and user.type() == "student":
                        return redirect('dashboard')
                    elif hasattr(user, 'type') and user.type() == "Faculty":
                        return redirect('faculty')
                    else:
                        return redirect('dashboard')
                except:
                    return redirect('dashboard')
        return render(request, 'login.html')
    except Exception as e:
        logging.error(f"Error in CustomLogin: {e}")
        return HttpResponse("An error occurred.")

@login_required
def CustomLogout(request):
    try:
        logout(request)
        return redirect('/')
    except Exception as e:
        logging.error(f"Error in CustomLogout: {e}")
        return HttpResponse("An error occurred.")

################################FillOut##########################
@login_required
def fillout(request):
    if request.user.is_superuser or request.user.type() == "Faculty":
        return redirect('forms/')
    else:
        return redirect('assigned/')

def create_form(request):
    if request.user.type() != "Faculty":
        return JsonResponse({"error": "Unauthorized"}, status=403)

    if request.method == "POST":
        try:
            title = request.POST.get("title")
            description = request.POST.get("description", "")
            assigned_ids = request.POST.get("assigned_students", "").split(",")

            form = FillOutForm.objects.create(
                title=title,
                description=description,
                created_by=request.user,
            )

            # Process fields
            fields = []
            for key in request.POST:
                if "field_name" in key:
                    idx = key.split("[")[1].split("]")[0]
                    field_name = request.POST.get(f"fields[{idx}][field_name]")
                    field_type = request.POST.get(f"fields[{idx}][field_type]")
                    options = request.POST.get(f"fields[{idx}][options]", None)
                    related_model = request.POST.get(f"fields[{idx}][related_model]", None)

                    fields.append(FillOutField(
                        form=form,
                        field_name=field_name,
                        field_type=field_type,
                        options=json.dumps([o.strip() for o in options.split(",")]) if options else None,
                        related_model=related_model if field_type == "file_awk" else None
                    ))

            FillOutField.objects.bulk_create(fields)
            form.assigned_students.set(student.objects.filter(id__in=assigned_ids))
            return JsonResponse({"message": "Form created", "form_id": form.id})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    print("Html Requested")
    all_students = student.objects.select_related("mentor").all()
    years = student.objects.values_list("year", flat=True).distinct()
    sections = student.objects.values_list("section", flat=True).distinct()

    return render(request, "create_form.html", {
        "students": all_students,
        "years": years,
        "sections": sections,
    })


def student_model_instances(request, model_name):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "Unauthorized"}, status=403)

    model_map = {
        "certificate": Certificate,
        "project": Projects,
    }

    model = model_map.get(model_name)
    if not model:
        return JsonResponse({"error": "Invalid model"}, status=400)

    instances = model.objects.filter(student=request.user)
    data = [{"id": obj.id, "label": str(obj)} for obj in instances]
    return JsonResponse(data, safe=False)


@csrf_exempt
def submit_response(request, form_id):
    if request.method == "POST":
        data = json.loads(request.body)
        form = get_object_or_404(FillOutForm, id=form_id)
        student_obj = request.user  # Assuming logged-in student
        FillOutResponse.objects.create(form=form, student=student_obj, responses=data["responses"])
        return JsonResponse({"message": "Response submitted successfully!"})
################################## Otp ##################################################
import smtplib
from email.mime.text import MIMEText
def send_otp(email):
    try:
        otp = random.randint(100000, 999999)
        logging.debug(f"Generated OTP: {otp}")
        sender_email = "webclubmits@gmail.com"
        sender_password = "zpco iaxk aywo rcwg"
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        recipient_email = email

        # Create message
        subject = "OTP for Verification"
        body = f"Your OTP for login into the Internship Dashboard is: {otp}"
        message = MIMEText(body)
        message["Subject"] = subject
        message["From"] = sender_email
        message["To"] = recipient_email

    # Connect to the server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            # Log in to the email account
            server.login(sender_email, sender_password)
            # Send the email
            server.sendmail(sender_email, recipient_email, message.as_string())
            print("OTP sent successfully.")
        return otp
    except Exception as e:
        logging.error(f"Error in send_otp: {e}")
        return None


def verify_otp(request):
    try:
        if request.method == 'POST':
            otp = request.POST.get('otp')
            print(otp)
            print(str(request.session['otp']))
            if otp == str(request.session['otp']):
                user = student.objects.create_user(
                    username=request.session.get('rollno'),
                    roll_no=request.session.get('rollno'),
                    is_staff=request.session.get('is_staff'),
                    first_name=request.session.get('first_name'),
                    dept=request.session.get('deptt'),
                    section=request.session.get('section'),
                    year=request.session.get('year'),
                    leetcode_user=request.session.get('leetcode_user'),
                )
                user.set_password(request.session.get('password'))
                user.save()
                print(f"{user.username}: {user.password}")
                leetcode = LeetCode.objects.create(rollno=user)
                leetcode.save()
                authenticated_user = authenticate(username=user.username, password=request.session.get('password'))
                if authenticated_user is not None:
                    login(request, authenticated_user, backend='flexapp.auth_backends.StudentBackend')
                    logging.debug(f"User logged in: {authenticated_user}")
                    return redirect('dashboard')
                else:
                    logging.error("Authentication failed.")
                    return redirect('login')
            else:
                logging.error('Invalid OTP')
                return render(request, 'verify_otp.html', {'error': 'Invalid OTP'})

        return render(request, 'verify_otp.html')
    except Exception as e:
        logging.error(f"Error in verify_otp: {e}")
        return HttpResponse("An error occurred.")
def register(request):
    try:
        if request.method == 'POST':
            request.session['rollno'] = request.POST.get('rollno')
            request.session['password'] = request.POST.get('password')
            request.session['first_name'] = request.POST.get('first_name')
            request.session['role'] = request.POST.get('role')
            request.session['deptt'] = request.POST.get('dept')
            request.session['section'] = request.POST.get('section')
            request.session['year'] = request.POST.get('year')
            request.session['is_staff'] = request.POST.get('role') == 'staff'
            request.session['leetcode_user'] = request.POST.get('leetcode_user')

            stu_email = f"{request.session['rollno']}@mits.ac.in"

            try:
                otp = send_otp(stu_email)
                request.session['otp'] = otp
                logging.debug(f"OTP sent: {otp}")

                return redirect('verify_otp')

            except Exception as e:
                logging.error(f"Error during registration: {e}")
                return render(request, 'register.html', {'error': str(e)})

        return render(request, 'register.html')
    except Exception as e:
        logging.error(f"Error in register: {e}")
        return HttpResponse("An error occurred.")
from django.db.models import Count, Avg
@login_required
def faculty(request):
    try:
        from .models import student, Certificate, Technology, Projects, Faculty

        # Get unique values for filter options
        certificate_providers = list(Certificate.objects.exclude(source='').values_list('source', flat=True).distinct().order_by('source'))
        
        # Combine both source into one list for the filter (removed course_provider as it doesn't exist)
        all_providers = sorted(list(set(certificate_providers)))
        
        # Get all technologies
        technologies = list(Technology.objects.all().values_list('name', flat=True).distinct().order_by('name'))
        
        # Get unique domains from certificates
        domains = list(Certificate.objects.exclude(domain='').values_list('domain', flat=True).distinct().order_by('domain'))
        
        # Get certificate categories
        certificate_categories = [choice[0] for choice in Certificate.CATEGORY_CHOICES]
        
        # Get project statuses
        project_statuses = [choice[0] for choice in Projects.status_choices]
        
        # Get faculty list for mentor filter
        mentors = list(Faculty.objects.all().values_list('first_name', flat=True).distinct().order_by('first_name'))
        
        # Get department choices
        departments = [choice[0] for choice in student.DEPT_CHOICES]

        studentData = list(
            student.objects.filter(is_superuser=False)
            .annotate(
                project_count=Count('projects', distinct=True),
                certificate_count=Count('certificates', distinct=True)
            )
            .values(
                'roll_no', 'first_name', 'dept', 'year', 'section',
                'studentrollno__TotalProblems',  # double-check this name
                'project_count', 'certificate_count', 'mentor__first_name'
            )
        )
        
        # Add additional information to each student
        for student_data in studentData:
            student_obj = student.objects.get(roll_no=student_data['roll_no'])
            
            # Certificate provider information
            student_certificates = Certificate.objects.filter(rollno=student_obj)
            student_providers = []
            student_domains = []
            student_cert_categories = []
            
            for cert in student_certificates:
                if cert.source:
                    student_providers.append(cert.source)
                if cert.domain:
                    student_domains.append(cert.domain)
                student_cert_categories.append(cert.category)
            
            student_data['certificate_providers'] = list(set(student_providers))
            student_data['domains'] = list(set(student_domains))
            student_data['certificate_categories'] = list(set(student_cert_categories))
            
            # Technology information from projects only (certificates don't have technologies field)
            student_technologies = []
            
            # Get technologies from projects
            student_projects = Projects.objects.filter(contributors=student_obj)
            for project in student_projects:
                project_techs = project.technologies.all().values_list('name', flat=True)
                student_technologies.extend(project_techs)
            
            student_data['technologies'] = list(set(student_technologies))
            
            # Project status information
            student_project_statuses = list(student_projects.values_list('status', flat=True))
            student_data['project_statuses'] = list(set(student_project_statuses))
            
        
        return render(request, 'faculty_dashboard.html', {
            'studentData': studentData,
            'certificate_providers': all_providers,
            'technologies': technologies,
            'domains': domains,
            'certificate_categories': certificate_categories,
            'project_statuses': project_statuses,
            'mentors': mentors,
            'departments': departments
        })

    except Exception as e:
        logging.error(f"Error in faculty: {e}")
        return HttpResponse("An error occurred.")
############################### DEMO #############################
from django.shortcuts import render, redirect
from .forms import CertificateForm
from .models import Certificate

def certificate_create(request):
    if request.method == 'POST':
        form = CertificateForm(request.POST, request.FILES)
        if form.is_valid():
            cert = form.save(commit=False)
            cert.rollno = request.user  # Assuming OneToOne link
            cert.save()
            form.save_m2m()
            return redirect('/')
    else:
        form = CertificateForm()

    context = {
        'form': form,
        'distinct_sources': Certificate.objects.values_list('source', flat=True).distinct(),
        'distinct_providers': Certificate.objects.values_list('course_provider', flat=True).distinct(),
        'distinct_domains': Certificate.objects.values_list('domain', flat=True).distinct(),
        'distinct_fests': Certificate.objects.values_list('fest_name', flat=True).distinct(),
    }

    return render(request, 'certificate_form.html', context)


################################ Coordinator #####################################
# views.py
from django.shortcuts import render
from .models import *
from django.contrib.auth.decorators import login_required

@login_required
def coordinator_dashboard(request):
    faculty = request.user  # assumes user is Faculty instance
    roles = faculty.coordinator_roles.all()

    # For layout rendering
    all_certificates = []
    all_projects = []
    all_publications = []

    for role in roles:
        if role.can_view_certificates:
            certs = Certificate.objects.all()
            if role.providers.exists():
                certs = certs.filter(provider__in=role.providers.values_list('name', flat=True))
            all_certificates.extend(list(certs))

        if role.can_view_projects:
            all_projects.extend(Projects.objects.all())

        if role.can_view_publications:
            all_publications.extend(publications.objects.all())

    context = {
        'roles': roles,
        'certificates': all_certificates,
        'projects': all_projects,
        'publications': all_publications
    }

    return render(request, 'dashboard/coordinator_dashboard.html', context)

################################ FLEX #####################################
@csrf_exempt  # Exempt from CSRF for cross-origin requests
def edit_project(request):
    try:
        if request.method == 'POST':
            # Check if we have JSON data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                primary_key = data.get('id')
                title = data.get('name')
                description = data.get('description')
                status = data.get('status')
                github_link = data.get('github_link')
                year_and_sem = data.get('year_and_sem')
                technologies = data.get('technologies')
                new_technologies = data.get('new_technologies', [])
                contributors = data.get('contributors', [])
            else:
                # Handle form data
                primary_key = request.POST.get('id')
                title = request.POST.get('name')
                description = request.POST.get('description')
                status = request.POST.get('status')
                github_link = request.POST.get('github_link')
                year_and_sem = request.POST.get('year_and_sem')
                technologies = request.POST.getlist('technologies')
                new_technologies = request.POST.getlist('new_technologies')
                contributors = request.POST.getlist('contributors')

            print(f"Editing project with ID: {primary_key}, Title: {title}, Description: {description}, Status: {status}, GitHub Link: {github_link}")

            project = Projects.objects.get(id=primary_key)
            project.title = title
            project.description = description
            project.status = status
            project.year_and_sem = year_and_sem
            project.github_link = github_link
            project.save()
            print(f"{technologies} saved")
            # Handle technologies if present
            if technologies:
                print(f"Selected technologies: {technologies}")
                for tech_id in technologies:
                    try:
                        tech = Technology.objects.get(id=tech_id)
                        project.technologies.add(tech)
                    except Technology.DoesNotExist:
                        pass
            if new_technologies != []:
                print(f"New technologies: {new_technologies}")
                for tech_name in new_technologies:
                    tech_name = tech_name.strip().capitalize()
                    if tech_name:
                        tech, created = Technology.objects.get_or_create(name=tech_name)
                        project.technologies.add(tech)

            print(f"Contributors: {contributors}")
            if contributors:
                project.contributors.clear()
                for contributor_id in contributors:
                    try:
                        contributor = student.objects.get(id=contributor_id)
                        project.contributors.add(contributor)
                    except student.DoesNotExist:
                        pass

            # Check the Accept header to decide the response format
            if 'application/json' in request.headers.get('Accept', ''):
                # Return JSON response for API requests
                return JsonResponse({'status': 'success', 'message': 'Project updated successfully'})
            else:
                # Return redirect for browser requests
                return redirect('dashboard')
    except Exception as e:
        logging.error(f"Error in edit_project: {e}")
        if 'application/json' in request.headers.get('Accept', ''):
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        else:
            return HttpResponse(f"An error occurred: {e}")

def edit_certification(request):
    try:
        if request.method == 'POST':
            primary_key = request.POST.get('id')
            source = request.POST.get('provider')
            title = request.POST.get('course-name')
            course_link = request.POST.get('course-link')
            project = Certificate.objects.get(id=primary_key)
            project.source = source
            project.title = title
            project.course_link = course_link
            project.save()
            return redirect('dashboard')
    except Exception as e:
        logging.error(f"Error in edit_certification: {e}")
        return HttpResponse("An error occurred.")

def delete_project(request, primary_key):
    try:
        project = Projects.objects.get(id=primary_key)
        project.delete()
        return JsonResponse({'status': 'success', 'message': 'Project deleted successfully', 'status_code': 200})
    except Exception as e:
        logging.error(f"Error in delete_project: {e}")
        return HttpResponse("An error occurred.")

def delete_certification(request, primary_key):
    try:
        project = Certificate.objects.get(id=primary_key)
        project.delete()
        return JsonResponse({'status': 'success', 'message': 'Certification deleted successfully', 'status_code': 200})
    except Exception as e:
        logging.error(f"Error in delete_certification: {e}")
        return HttpResponse("An error occurred.")

def leetcode_request(request, leetcode_user):
    try:
        username = leetcode_user
        query = '''
        {
            matchedUser(username: "%s") {
                username
                submitStats: submitStatsGlobal {
                    acSubmissionNum {
                        difficulty
                        count
                        submissions
                    }
                }
            }
        }
        ''' % username

        url = 'https://leetcode.com/graphql'
        response = requests.post(url, json={'query': query})

        if response.status_code == 200:
            data = response.json()
            submission_data = data['data']['matchedUser']['submitStats']['acSubmissionNum']

            easy_count = next((item['count'] for item in submission_data if item['difficulty'] == 'Easy'), 0)
            medium_count = next((item['count'] for item in submission_data if item['difficulty'] == 'Medium'), 0)
            hard_count = next((item['count'] for item in submission_data if item['difficulty'] == 'Hard'), 0)
        else:
            easy_count = medium_count = hard_count = 0
        leetcode_user, created = LeetCode.objects.get_or_create(rollno=student.objects.get(leetcode_user=username))
        leetcode_user.easy = easy_count
        leetcode_user.medium = medium_count
        leetcode_user.hard = hard_count
        leetcode_user.TotalProblems = int(easy_count) + int(medium_count) + int(hard_count)
        leetcode_user.save()
        return JsonResponse({
            'easy_count': easy_count,
            'medium_count': medium_count,
            'hard_count': hard_count,
        })
    except Exception as e:
        logging.error(f"Error in leetcode_request: {e}")
        return JsonResponse({'error': 'An error occurred.'}, status=500)

################################ Download #####################################

def download_request(request):
    try:
        list = request.body
        list = json.loads(list)
        students = student.objects.filter(roll_no__in=list)
        data = []
        for stu in students:
            leetcode = LeetCode.objects.get(rollno=stu)
            technical = Certificate.objects.filter(rollno=stu, category="Technical")
            foreign = Certificate.objects.filter(rollno=stu, category="Foreign Language")
            projects = Projects.objects.filter(rollno=stu)
            data.append({
                'Roll No': stu.roll_no,
                'Student Name': stu.first_name,
                'Department': stu.dept,
                'Section': stu.section,
                'Year': stu.year,
                'Total Count': leetcode.TotalProblems,
                'Projects': ",".join([project.title for project in projects]),
                'Foreign Languages': ",".join([project.title for project in foreign]),
                'Technical Certificates': ",".join([project.title for project in technical]),
            })

        df = pd.DataFrame(data)

        # Create a response object and specify the content type for Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'

        # Write the DataFrame to the response object as an Excel file
        df.to_excel(response, index=False)

        return response
    except Exception as e:
        logging.error(f"Error in download_request: {e}")
        return HttpResponse("An error occurred.")

def studentView(request, rollno):
    try:
        if request.user.type() == "Faculty":
            context = getStudentDetails(student.objects.get(roll_no=rollno))
            return render(request, 'dashboard.html', context=context)
        return HttpResponse("<h1>You are not authorized to view this page. Redirecting...</h1>\n<script>setTimeout( () => {window.location.href='/';}, 2000)</script>")
    except Exception as e:
        logging.error(f"Error in studentView: {e}")
        return HttpResponse("An error occurred.")

def forgot_password(request):
    try:
        if request.method == 'POST':
            rollno = request.POST.get('rollno')
            user = student.objects.get(roll_no=rollno)
            otp = send_otp(f"{user.roll_no}@mits.ac.in")
            print(otp)
            request.session['rollno'] = rollno
            request.session['otp'] = otp
            print(request.session['otp'])
            return redirect('verify_otp_forgot')
        else:
            return render(request, 'forgot_password.html')
    except Exception as e:
        logging.error(f"Error in forgot_password: {e}")
        return HttpResponse("An error occurred.")

def verify_otp_forgot(request):
    try:
        if request.method == 'POST':
            otp = request.POST.get('otp')
            print(otp)
            print(str(request.session['otp']))
            if otp == str(request.session['otp']):
                print("OTP Verified")
                return redirect('reset_password')
            else:
                return render(request, 'verify_otp.html', {'error': 'Invalid OTP'})
        return render(request, 'verify_otp.html')
    except Exception as e:
        logging.error(f"Error in verify_otp_forgot: {e}")
        return HttpResponse("An error occurred.")

def reset_password(request):
    if request.method == 'POST':
        password = request.POST.get('confirm_password')
        user = student.objects.get(roll_no=request.session['rollno'])
        print(user)
        print(password)
        user.set_password(password)
        user.save()
        return redirect('login')
    return render(request, 'reset_password.html')



@login_required
def student_profile(request):
    user = request.user  # Get the currently logged-in user

    if request.method == "POST":
        # Track which fields were updated
        updated_fields = {}

        # Basic Information Fields
        if 'name' in request.POST:
            name = request.POST.get("name")
            if name and name != user.first_name:
                user.first_name = name
                updated_fields['name'] = name

        if 'username' in request.POST:
            username = request.POST.get("username")
            if username and username != user.username:
                user.username = username
                updated_fields['username'] = username

        if 'personal_email' in request.POST:
            personal_email = request.POST.get("personal_email")
            if personal_email != user.personal_email:
                user.personal_email = personal_email or None
                updated_fields['personal_email'] = personal_email

        if 'phone' in request.POST:
            phone = request.POST.get("phone")
            if phone != user.phone:
                user.phone = phone or None
                updated_fields['phone'] = phone

        if 'date_of_birth' in request.POST:
            date_of_birth = request.POST.get("date_of_birth")
            if date_of_birth and date_of_birth != str(user.date_of_birth):
                try:
                    from datetime import datetime
                    user.date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
                    updated_fields['date_of_birth'] = date_of_birth
                except ValueError:
                    pass  # Invalid date format, skip

        # Academic Information Fields
        if 'roll_no' in request.POST:
            roll_no = request.POST.get("roll_no")
            if roll_no and roll_no != user.roll_no:
                user.roll_no = roll_no
                updated_fields['roll_no'] = roll_no

        if 'dept' in request.POST:
            dept = request.POST.get("dept")
            if dept and dept != user.dept:
                user.dept = dept
                updated_fields['dept'] = dept

        if 'year' in request.POST:
            year = request.POST.get("year")
            if year and str(year) != str(user.year):
                user.year = int(year)
                updated_fields['year'] = year

        if 'section' in request.POST:
            section = request.POST.get("section")
            if section and section != user.section:
                user.section = section
                updated_fields['section'] = section

        if 'current_cgpa' in request.POST:
            current_cgpa = request.POST.get("current_cgpa")
            if current_cgpa:
                try:
                    cgpa_value = float(current_cgpa)
                    if 0 <= cgpa_value <= 10:
                        user.current_cgpa = cgpa_value
                        updated_fields['current_cgpa'] = current_cgpa
                except ValueError:
                    pass  # Invalid CGPA format, skip

        if 'admission_year' in request.POST:
            admission_year = request.POST.get("admission_year")
            if admission_year:
                try:
                    year_value = int(admission_year)
                    user.admission_year = year_value
                    updated_fields['admission_year'] = admission_year
                except ValueError:
                    pass  # Invalid year format, skip

        # Professional Links
        if 'leetcode_user' in request.POST:
            leetcode_user = request.POST.get("leetcode_user")
            if leetcode_user != user.leetcode_user:
                user.leetcode_user = leetcode_user or "Username"
                updated_fields['leetcode_user'] = leetcode_user

        if 'githublink' in request.POST:
            githublink = request.POST.get("githublink")
            if githublink != str(user.githublink or ''):
                user.githublink = githublink or None
                updated_fields['githublink'] = githublink

        if 'linkedin_url' in request.POST:
            linkedin_url = request.POST.get("linkedin_url")
            if linkedin_url != str(user.linkedin_url or ''):
                user.linkedin_url = linkedin_url or None
                updated_fields['linkedin_url'] = linkedin_url

        if 'portfolio_url' in request.POST:
            portfolio_url = request.POST.get("portfolio_url")
            if portfolio_url != str(user.portfolio_url or ''):
                user.portfolio_url = portfolio_url or None
                updated_fields['portfolio_url'] = portfolio_url

        # Skills and Career Information
        if 'technical_skills' in request.POST:
            technical_skills = request.POST.get("technical_skills")
            if technical_skills != str(user.technical_skills or ''):
                user.technical_skills = technical_skills or None
                updated_fields['technical_skills'] = technical_skills

        if 'soft_skills' in request.POST:
            soft_skills = request.POST.get("soft_skills")
            if soft_skills != str(user.soft_skills or ''):
                user.soft_skills = soft_skills or None
                updated_fields['soft_skills'] = soft_skills

        if 'career_interests' in request.POST:
            career_interests = request.POST.get("career_interests")
            if career_interests != str(user.career_interests or ''):
                user.career_interests = career_interests or None
                updated_fields['career_interests'] = career_interests

        # Guardian Information
        if 'guardian_name' in request.POST:
            guardian_name = request.POST.get("guardian_name")
            if guardian_name != str(user.guardian_name or ''):
                user.guardian_name = guardian_name or None
                updated_fields['guardian_name'] = guardian_name

        if 'guardian_phone' in request.POST:
            guardian_phone = request.POST.get("guardian_phone")
            if guardian_phone != str(user.guardian_phone or ''):
                user.guardian_phone = guardian_phone or None
                updated_fields['guardian_phone'] = guardian_phone

        if 'address' in request.POST:
            address = request.POST.get("address")
            if address != str(user.address or ''):
                user.address = address or None
                updated_fields['address'] = address

        # Save changes if any fields were updated
        if updated_fields:
            try:
                # Update profile completion percentage
                total_fields = 20  # Total number of profile fields
                filled_fields = sum([
                    1 for field in [
                        user.first_name, user.username, user.email, user.personal_email,
                        user.phone, user.date_of_birth, user.roll_no, user.dept,
                        user.year, user.section, user.current_cgpa, user.admission_year,
                        user.leetcode_user, user.githublink, user.linkedin_url, user.portfolio_url,
                        user.technical_skills, user.career_interests, user.guardian_name, user.address
                    ] if field and str(field).strip() not in ['', 'Username']
                ])
                user.profile_completion_percentage = int((filled_fields / total_fields) * 100)
                
                user.save()
                messages.success(request, f"Profile updated successfully! {len(updated_fields)} field(s) changed.")
            except Exception as e:
                messages.error(request, f"An error occurred while saving: {e}")
        else:
            messages.info(request, "No changes were made to your profile.")

        return redirect('profile')
    
    return render(request, 'student_profile_edit.html', {'user': user})


def change_ac(request,yearr):
    if request.user.is_authenticated and request.user.type() == "Faculty":
        studs = student.objects.all()
        for stu in studs:
            stu.year += yearr
            stu.save()
        return redirect('faculty')

@csrf_exempt
def upload_students(request):
    if request.method == 'POST' and request.FILES['file']:
        try:
            file = request.FILES['file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            file_path = fs.path(filename)

            # Read the Excel file
            df = pd.read_excel(file_path)

            # Iterate through the rows and add details to the student model
            for index, row in df.iterrows():
                s = student(
                    roll_no=row['Roll'],
                    first_name=row['Name'],
                    year=row['Year'],
                    section=row['Section'],
                    username=row['Roll']
                )
                s.set_password(row['Roll'])
                s.save()

            messages.success(request, "Students added successfully.")
            return redirect('dashboard')
        except Exception as e:
            logging.error(f"Error in upload_students: {e}")
            return HttpResponse("An error occurred.")
    return render(request, 'upload_students.html')


def placement_coordinator_required(view_func):
    """Decorator that requires user to be a placement coordinator"""
    @login_required
    def wrapper(request, *args, **kwargs):
        if not is_placement_coordinator(request.user):
            from django.shortcuts import render
            return render(request, '403.html', status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

######################################################################################
@placement_coordinator_required
def placement_dashboard(request):
    year = request.GET.get('year')
    tech = request.GET.get('tech')
    cert_cat = request.GET.get('cert_cat')
    domain = request.GET.get('domain')

    students = student.objects.all()
    leetcode_data = LeetCode.objects.all()
    cert_data = Certificate.objects.all()
    proj_data = Projects.objects.all()

    # if year:
    #     students = students.filter(year_and_sem__icontains=year)
    #     leetcode_data = leetcode_data.filter(rollno__year_and_sem__icontains=year)
    #     cert_data = cert_data.filter(rollno__year_and_sem__icontains=year)
    #     proj_data = proj_data.filter(year_and_sem__icontains=year)

    if tech:
        proj_data = proj_data.filter(technologies__name__icontains=tech)

    if cert_cat:
        cert_data = cert_data.filter(category=cert_cat)

    if domain:
        proj_data = proj_data.filter(domain__icontains=domain)

    context = {
        'leetcode_data': list(leetcode_data.values('rollno__dept').annotate(avg_total=Avg('TotalProblems'))),
        'cert_data': list(cert_data.values('rollno__dept', 'category').annotate(count=Count('id'))),
        'proj_data': list(proj_data.values('year_and_sem').annotate(total=Count('id'))),
        'filters': {
            'selected_year': year,
            'selected_tech': tech,
            'selected_cert_cat': cert_cat,
            'selected_domain': domain,
        }
    }
    return render(request, 'dashboard/placement_dashboard.html', context)

# views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import FillOutForm, FillOutField, student
from django.contrib.auth.decorators import login_required
import json


@login_required
def list_assigned_forms(request):
    assigned_forms = request.user.assigned_forms.all()
    return render(request, "fill_form_list.html", {"assigned_forms": assigned_forms})

from .models import FillOutResponse
@login_required
def fill_form(request, form_id):
    form = FillOutForm.objects.get(id=form_id)
    student = request.user

    certificates = Certificate.objects.filter(rollno=student)

    if request.method == "POST":
        responses = {}
        for field in form.fields.all():
            responses[field.field_name] = request.POST.get(str(field.id))
        FillOutResponse.objects.create(form=form, rollno=student, responses=responses)
        return redirect("assigned/")  # or wherever you want

    return render(request, "fill_form.html", {
        "form": form,
        "certificates": certificates
    })

from django.shortcuts import render, get_object_or_404, redirect
from .models import FillOutForm, FillOutField, Certificate

def get_form(request, form_id):
    form = get_object_or_404(FillOutForm, id=form_id)
    print(form)
    student = request.user
    certificates = Certificate.objects.filter(rollno=student)
    print(certificates)

    if request.method == "POST":
        responses = {}
        for field in form.fields.all():
            responses[field.field_name] = request.POST.get(str(field.id))
        form_response = FillOutResponse.objects.create(
            form=form,
            student=student,
            responses=responses
        )
        return redirect("dashboard")

    return render(request, "fill_form_detail.html", {
        "form": form,
        "fields": form.fields.all(),
        "model_instances": {
        "certificate": Certificate.objects.filter(rollno=request.user),
        "project": Projects.objects.filter(contributors=request.user),}
        })

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from .models import FillOutForm, FillOutResponse, student
import csv

def form_list_view(request):
    faculty = request.user
    forms = FillOutForm.objects.filter(created_by=faculty)
    return render(request, 'dashboard/form_list.html', {'forms': forms})


from django.apps import apps
from django.shortcuts import render, get_object_or_404
from .models import FillOutForm, FillOutResponse

def form_detail_view(request, form_id):
    form = get_object_or_404(FillOutForm, id=form_id)
    responses = FillOutResponse.objects.filter(form=form).select_related("student")

    model_instance_map = {}

    for response in responses:
        for field in form.fields.filter(field_type="file_awk"):
            field_name = field.field_name
            model_name = field_name  # Assuming model is named same as field

            id_value = response.responses.get(field_name)
            if id_value:
                try:
                    model = apps.get_model("flexapp", model_name)
                    instance = model.objects.get(id=id_value)
                    model_instance_map[(response.id, field_name)] = instance
                except Exception:
                    model_instance_map[(response.id, field_name)] = None

    context = {
        "form": form,
        "responses": responses,
        "model_instance_map": model_instance_map,
    }
    return render(request, "dashboard/form_detail.html", context)


def download_csv(request, form_id, download_type):
    form = get_object_or_404(FillOutForm, id=form_id)
    response = HttpResponse(content_type='text/csv')

    if download_type == "responses":
        filename = f"{form.title}_responses.csv"
        writer = csv.writer(response)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer.writerow(["Student Roll No", "Student Name", "Submitted At"] + [f.field_name for f in form.fields.all()])
        for r in FillOutResponse.objects.filter(form=form):
            student_info = [r.student.roll_no, r.student.first_name, r.submitted_at]
            answers = [r.responses.get(f.field_name, "") for f in form.fields.all()]
            writer.writerow(student_info + answers)

    elif download_type == "filled":
        filename = f"{form.title}_filled.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Roll No", "Name"])
        for student_obj in form.assigned_students.filter(id__in=FillOutResponse.objects.filter(form=form).values_list("student_id", flat=True)):
            writer.writerow([student_obj.roll_no, student_obj.first_name])

    elif download_type == "not_filled":
        filename = f"{form.title}_not_filled.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(["Roll No", "Name"])
        for student_obj in form.assigned_students.exclude(id__in=FillOutResponse.objects.filter(form=form).values_list("student_id", flat=True)):
            writer.writerow([student_obj.roll_no, student_obj.first_name])

    return response

@api_view(['GET'])
@permission_classes([AllowAny])
def project_detail(request, project_id):
    try:
        project = get_object_or_404(Projects, id=project_id)
        serializer = ProjectSerializer(project, context={'request': request})

        # Include additional details for editing
        response_data = serializer.data

        # Add technologies
        response_data['technologies'] = []
        for tech in project.technologies.all():
            response_data['technologies'].append({
                'id': tech.id,
                'name': tech.name
            })

        # Add contributors
        response_data['contributors'] = []
        for contributor in project.contributors.all():
            response_data['contributors'].append({
                'id': contributor.id,
                'rollno': contributor.roll_no,
                'name': contributor.first_name,
                'last_name': contributor.last_name
            })

        return Response(response_data)
    except Exception as e:
        logging.error(f"Error in project_detail: {e}")
        return Response({'error': str(e)}, status=500)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    users = User.objects.all()
    return render(request, 'admin_dashboard.html', {'users': users, 'user': request.user})

@user_passes_test(lambda u: u.is_superuser)
def admin_create_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        is_superuser = 'is_superuser' in request.POST
        user = User(username=username, email=email, is_superuser=is_superuser, is_staff=is_superuser)
        user.set_password(password)
        user.save()
        messages.success(request, 'User created successfully!')
        return redirect('admin_dashboard')
    return redirect('admin_dashboard')

@user_passes_test(lambda u: u.is_superuser)
def admin_update_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.email = request.POST.get('email', user.email)
        user.is_superuser = 'is_superuser' in request.POST
        user.is_staff = user.is_superuser
        if request.POST.get('password'):
            user.set_password(request.POST['password'])
        user.save()
        messages.success(request, 'User updated successfully!')
        return redirect('admin_dashboard')
    return render(request, 'admin_update_user.html', {'user_obj': user})

@user_passes_test(lambda u: u.is_superuser)
def admin_delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully!')
        return redirect('admin_dashboard')
    return redirect('admin_dashboard')

from django.http import HttpResponseBadRequest
@user_passes_test(lambda u: u.is_superuser)
def admin_create_student(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        roll_no = request.POST.get('roll_no', '')
        year = request.POST.get('year', 3)
        section = request.POST.get('section', 'A')
        dept = request.POST.get('dept', 'CSE')
        s = student(username=username, email=email, roll_no=roll_no, year=year, section=section, dept=dept)
        s.set_password(password)
        s.save()
        messages.success(request, 'Student created successfully!')
        return redirect('admin_dashboard')
    return HttpResponseBadRequest()
@user_passes_test(lambda u: u.is_superuser)
def admin_update_student(request, student_id):
    s = get_object_or_404(student, id=student_id)
    if request.method == 'POST':
        s.email = request.POST.get('email', s.email)
        s.roll_no = request.POST.get('roll_no', s.roll_no)
        s.year = request.POST.get('year', s.year)
        s.section = request.POST.get('section', s.section)
        s.dept = request.POST.get('dept', s.dept)
        if request.POST.get('password'):
            s.set_password(request.POST['password'])
        s.save()
        messages.success(request, 'Student updated successfully!')
        return redirect('admin_dashboard')
    return render(request, 'admin_update_user.html', {'user_obj': s, 'is_student': True})
@user_passes_test(lambda u: u.is_superuser)
def admin_delete_student(request, student_id):
    s = get_object_or_404(student, id=student_id)
    if request.method == 'POST':
        s.delete()
        messages.success(request, 'Student deleted successfully!')
        return redirect('admin_dashboard')
    return HttpResponseBadRequest()
@user_passes_test(lambda u: u.is_superuser)
def admin_create_faculty(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        dept = request.POST.get('dept', 'CSE')
        default_year = request.POST.get('default_year', None)
        default_section = request.POST.get('default_section', 'A')
        f = Faculty(username=username, email=email, dept=dept, default_year=default_year, default_section=default_section)
        f.set_password(password)
        f.save()
        messages.success(request, 'Faculty created successfully!')
        return redirect('admin_dashboard')
    return HttpResponseBadRequest()
@user_passes_test(lambda u: u.is_superuser)
def admin_update_faculty(request, faculty_id):
    f = get_object_or_404(Faculty, id=faculty_id)
    if request.method == 'POST':
        f.email = request.POST.get('email', f.email)
        f.dept = request.POST.get('dept', f.dept)
        f.default_year = request.POST.get('default_year', f.default_year)
        f.default_section = request.POST.get('default_section', f.default_section)
        if request.POST.get('password'):
            f.set_password(request.POST['password'])
        f.save()
        messages.success(request, 'Faculty updated successfully!')
        return redirect('admin_dashboard')
    return render(request, 'admin_update_user.html', {'user_obj': f, 'is_faculty': True})
@user_passes_test(lambda u: u.is_superuser)
def admin_delete_faculty(request, faculty_id):
    f = get_object_or_404(Faculty, id=faculty_id)
    if request.method == 'POST':
        f.delete()
        messages.success(request, 'Faculty deleted successfully!')
        return redirect('admin_dashboard')
    return HttpResponseBadRequest()
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from .models import student, Faculty
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    students = student.objects.all()
    faculty = Faculty.objects.all()
    return render(request, 'admin_dashboard.html', {
        'students': students,
        'faculty': faculty,
        'user': request.user
    })

@user_passes_test(lambda u: u.is_superuser)
def admin_create_user(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        is_superuser = 'is_superuser' in request.POST
        user = User(username=username, email=email, is_superuser=is_superuser, is_staff=is_superuser)
        user.set_password(password)
        user.save()
        messages.success(request, 'User created successfully!')
        return redirect('admin_dashboard')
    return redirect('admin_dashboard')

@user_passes_test(lambda u: u.is_superuser)
def admin_update_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.email = request.POST.get('email', user.email)
        user.is_superuser = 'is_superuser' in request.POST
        user.is_staff = user.is_superuser
        if request.POST.get('password'):
            user.set_password(request.POST['password'])
        user.save()
        messages.success(request, 'User updated successfully!')
        return redirect('admin_dashboard')
    return render(request, 'admin_update_user.html', {'user_obj': user})

@user_passes_test(lambda u: u.is_superuser)
def admin_delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully!')
        return redirect('admin_dashboard')
    return redirect('admin_dashboard')


from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.http import JsonResponse
from .models import student, Certificate, Projects, PlacementOffer, LeetCode
import requests
import os
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied

def is_coordinator(user):
    return hasattr(user, 'coordinatorrole') and user.coordinatorrole is not None

def is_placement_coordinator(user):
    """Check if user is a placement coordinator with placement viewing permissions"""
    if not user.is_authenticated:
        return False
    
    if user.is_superuser:
        return True
    
    # Check if user is faculty and has placement coordinator role
    try:
        if hasattr(user, 'coordinator_roles'):
            for role in user.coordinator_roles.all():
                if role.can_view_placement:
                    return True
    except:
        pass
    
    return False

# FlexOn Dashboard Implementation
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Q
import traceback

def generate_orm_from_query(query):
    """
    Advanced query-to-ORM mapping using Google Gemini AI.
    Converts natural language queries to Django ORM strings with better accuracy.
    """
    import google.generativeai as genai
    import os
    from django.conf import settings
    
    try:
        # Configure Gemini AI
        genai.configure(api_key=getattr(settings, 'GEMINI_API_KEY', os.environ.get('GEMINI_API_KEY')))
        model = genai.GenerativeModel('gemini-pro')
        
        # Database schema context for better ORM generation
        schema_context = """
        Available Django Models and their fields:
        
        1. student (Main student model):
           - roll_no, name, email, dept, year, sem, phone, mentor
           - Related: projects, certificate, leetcode, placementoffer
        
        2. Projects:
           - rollno (FK to student), title, description, github_link, year, sem
           - Related: technologies (M2M with Technology)
        
        3. Certificate:
           - rollno (FK to student), title, source, course_provider, category, year, sem
        
        4. LeetCode:
           - rollno (FK to student), TotalProblems, Easy, Medium, Hard
        
        5. Technology:
           - name (unique technology name like 'Python', 'Java', etc.)
        
        6. Placement:
           - rollno (FK to student), company, package, status
        
        7. PlacementOffer:
           - rollno (FK to student), company, package, offer_type
        
        8. Faculty:
           - name, email, department, designation
        
        Common Django ORM patterns:
        - Filtering: Model.objects.filter(field__condition=value)
        - Annotations: Model.objects.annotate(alias=Count/Sum/Avg('related_field'))
        - Ordering: .order_by('field') or .order_by('-field') for descending
        - Relationships: Use double underscore __ for foreign keys
        - Aggregation functions: Count, Sum, Avg, Max, Min
        - Q objects for complex queries: Q(condition1) & Q(condition2)
        
        Important: Only generate READ-ONLY ORM queries. No create, update, delete operations.
        """
        
        prompt = f"""
        {schema_context}
        
        Convert this natural language query to Django ORM code:
        "{query}"
        
        Requirements:
        1. Return ONLY the ORM query string, no explanations
        2. Use proper Django ORM syntax
        3. Handle relationships correctly with double underscores
        4. Apply appropriate filters, annotations, and ordering
        5. Limit results to 50 records maximum using [:50] if needed
        6. Use select_related() for foreign keys when appropriate
        7. Use prefetch_related() for many-to-many relationships when needed
        
        Example outputs:
        - "student.objects.filter(dept='CSE').order_by('roll_no')[:50]"
        - "LeetCode.objects.select_related('rollno').order_by('-TotalProblems')[:10]"
        - "Projects.objects.prefetch_related('technologies').filter(technologies__name__icontains='python')"
        
        Query: {query}
        ORM:
        """
        
        response = model.generate_content(prompt)
        orm_query = response.text.strip()
        
        # Clean up the response - remove any markdown formatting or extra text
        if '```' in orm_query:
            import re
            code_blocks = re.findall(r'```(?:python)?\s*(.*?)\s*```', orm_query, re.DOTALL)
            if code_blocks:
                orm_query = code_blocks[0].strip()
        
        # Remove any leading/trailing quotes
        orm_query = orm_query.strip('"\'')
        
        # Validate that it's a proper ORM query
        if not orm_query or not any(model in orm_query for model in ['student.objects', 'Projects.objects', 'Certificate.objects', 'LeetCode.objects', 'Placement.objects', 'Technology.objects', 'Faculty.objects']):
            raise ValueError("Invalid ORM query generated")
        
        return orm_query
        
    except Exception as e:
        print(f"Gemini AI error: {e}")
        # Fallback to basic keyword matching if Gemini fails
        return _generate_fallback_orm(query)

def _generate_fallback_orm(query):
    """
    Fallback query generation when Gemini AI is unavailable.
    """
    query_lower = query.lower()
    
    # Basic query patterns
    if 'top' in query_lower and 'leetcode' in query_lower:
        return "LeetCode.objects.select_related('rollno').order_by('-TotalProblems')[:10]"
    
    elif 'students' in query_lower and 'python' in query_lower and 'projects' in query_lower:
        return "student.objects.annotate(python_projects=Count('projects', filter=Q(projects__technologies__name__icontains='python'))).filter(python_projects__gt=0)[:50]"
    
    elif 'placement' in query_lower and 'offers' in query_lower:
        return "student.objects.annotate(offer_count=Count('placementoffer')).filter(offer_count__gt=0)[:50]"
    
    elif 'cloud' in query_lower and ('cert' in query_lower or 'certificate' in query_lower):
        return "Certificate.objects.filter(Q(source__icontains='cloud') | Q(title__icontains='cloud'))[:50]"
    
    elif 'placement' in query_lower and 'stats' in query_lower:
        return "student.objects.values('dept').annotate(total=Count('id'), placed=Count('placementoffer')).order_by('dept')"
    
    elif 'students' in query_lower and 'projects' in query_lower:
        return "student.objects.annotate(project_count=Count('projects')).filter(project_count__gt=0)[:50]"
    
    elif 'certificates' in query_lower:
        return "Certificate.objects.select_related('rollno').order_by('-id')[:50]"
    
    elif 'leetcode' in query_lower:
        return "LeetCode.objects.select_related('rollno').order_by('-TotalProblems')[:50]"
    
    elif 'projects' in query_lower:
        return "Projects.objects.select_related('rollno').prefetch_related('technologies').order_by('-id')[:50]"
    
    elif 'students' in query_lower:
        return "student.objects.all().order_by('roll_no')[:50]"
    
    else:
        return "student.objects.all()[:20]"


def _normalize_orm_line(orm_line: str, model_aliases: dict) -> str:
    """
    Rewrite '<Alias>.objects' -> '<actual>.objects' using ONLY the provided model_aliases.
    Adds singular/plural + case variants at runtime without mutating model_aliases.
    """
    if not orm_line:
        return orm_line

    import re
    
    # Build an expanded lookup without changing the original dict
    expanded = {}
    for alias, actual in model_aliases.items():
        variants = {alias, alias.lower(), alias.capitalize()}

        # add singular/plural variants
        if alias.endswith('s'):
            singular = alias[:-1]
            variants.update({singular, singular.lower(), singular.capitalize()})
        else:
            plural = alias + 's'
            variants.update({plural, plural.lower(), plural.capitalize()})

        for v in variants:
            expanded[v] = actual  # all map to the same "actual"

    # Replace ONLY tokens that are used as '<name>.objects'
    pattern = re.compile(r"\b([A-Za-z_][A-Za-z0-9_]*)\s*\.objects")

    def repl(m):
        name = m.group(1)
        actual = expanded.get(name) or expanded.get(name.lower())
        return f"{actual}.objects" if actual else m.group(0)

    return pattern.sub(repl, orm_line)


@csrf_exempt
def flexon_dashboard(request):
    if request.method == 'POST':
        query = request.POST.get('query', '').strip()
        clarification = ''
        results = []
        columns = []

        # Step 1: Generate ORM from query (returns a string)
        orm_line = generate_orm_from_query(query)

        # Step 2: Normalize model references using the GIVEN model_aliases (unchanged)
        model_aliases = {
            'Student': 'student',
            'Faculty': 'Faculty',
            'Projects': 'Projects',
            'Certificate': 'Certificate',
            'LeetCode': 'LeetCode',
            'Placement': 'Placement',
            'Technology': 'Technology',
        }
        orm_line = _normalize_orm_line(orm_line, model_aliases) if orm_line else None

        # Step 3: Safety check (read-only)
        unsafe_keywords = [
            'create', 'update', 'delete', 'save',
            'bulk_create', 'bulk_update', 'remove', 'clear', 'set'
        ]
        if not orm_line:
            clarification = "Couldn't generate a valid ORM query."
        elif any(kw in orm_line.lower() for kw in unsafe_keywords):
            clarification = "Unsafe ORM operation detected. Only read-only queries are allowed."
        else:
            try:
                # Step 4: Evaluate safely
                from .models import Faculty, Technology, Placement
                from django.db import models as django_models
                
                safe_globals = {
                    '__builtins__': {},  # lock down builtins
                    # models/modules
                    'models': django_models,
                    'Count': Count,
                    'Q': Q,
                    # model classes available to the ORM line
                    'student': student,
                    'Faculty': Faculty,
                    'Projects': Projects,
                    'Certificate': Certificate,
                    'LeetCode': LeetCode,
                    'Placement': Placement,
                    'Technology': Technology,
                    'PlacementOffer': PlacementOffer,
                }
                print(f"Executing ORM: {orm_line}")

                queryset = eval(orm_line, safe_globals, {})

                # Step 5: Normalize results to list-of-dicts and build columns
                from django.db.models import QuerySet
                
                def as_dicts_from_instances(qs):
                    model = qs.model
                    field_names = [f.name for f in model._meta.concrete_fields]
                    return list(qs.values(*field_names))

                if isinstance(queryset, dict):
                    # Aggregate: {'count': X, ...}
                    results = [queryset]
                    columns = list(queryset.keys())
                elif isinstance(queryset, (list, tuple)):
                    if queryset and isinstance(queryset[0], tuple):
                        # values_list / tuples -> map to col_1, col_2...
                        col_names = [f"col_{i+1}" for i in range(len(queryset[0]))]
                        results = [dict(zip(col_names, row)) for row in queryset[:50]]
                        columns = col_names
                    elif queryset and hasattr(queryset[0], '__class__'):
                        # list of model instances
                        model = queryset[0].__class__
                        field_names = [f.name for f in model._meta.concrete_fields]
                        results = [ {fn: getattr(obj, fn) for fn in field_names} for obj in queryset[:50] ]
                        columns = field_names
                    else:
                        results = []
                        columns = []
                elif isinstance(queryset, QuerySet):
                    # If caller already used .values(), don't double-process
                    try:
                        sample = queryset[:1]
                        # If it's a ValuesQuerySet, converting to list gives dicts
                        if sample and isinstance(list(sample)[0], dict):
                            results = list(queryset)[:50]
                        else:
                            results = as_dicts_from_instances(queryset)[:50]
                    except Exception:
                        # fallback to values() anyway
                        results = list(queryset.values())[:50]

                    columns = list(results[0].keys()) if results else []
                else:
                    # Any other scalar/object -> wrap it
                    results = [{"result": queryset}]
                    columns = ["result"]

            except Exception as e:
                clarification = f"Error executing ORM: {e}"
                traceback.print_exc()

        return JsonResponse({
            'clarification': clarification,
            'results': results,
            'columns': columns,
            'chart': None
        })

    return render(request, 'flexon_dashboard.html')


########################### NAAC Report Generation ###########################
from django.views.decorators.http import require_POST

@csrf_exempt
@require_POST
def generate_naac_report(request):
    """Generate comprehensive NAAC compliance report for students"""
    try:
        # Parse request data
        data = json.loads(request.body)
        selected_students = data.get('students', [])
        include_all = data.get('include_all', False)
        
        # Get student queryset
        if include_all or not selected_students:
            students_queryset = student.objects.filter(is_superuser=False)
        else:
            students_queryset = student.objects.filter(roll_no__in=selected_students)
        
        # Verify we have students to process
        if not students_queryset.exists():
            return JsonResponse({'error': 'No students found to generate report'}, status=400)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # Generate basic student summary sheet first - this is essential
        try:
            generate_student_summary_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Critical error in student summary: {e}")
            # Create a simple fallback sheet
            ws = wb.create_sheet(title="Student Data")
            ws.cell(row=1, column=1, value="Roll Number")
            ws.cell(row=1, column=2, value="Name")
            ws.cell(row=1, column=3, value="Department")
            
            for row, student_obj in enumerate(students_queryset, 2):
                ws.cell(row=row, column=1, value=getattr(student_obj, 'roll_no', 'N/A'))
                ws.cell(row=row, column=2, value=f"{getattr(student_obj, 'first_name', '')} {getattr(student_obj, 'last_name', '')}".strip())
                ws.cell(row=row, column=3, value=getattr(student_obj, 'dept', 'N/A'))
        
        # Try to generate other sheets, but don't fail if there are issues
        try:
            generate_academic_performance_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating academic performance sheet: {e}")
        
        try:
            generate_projects_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating projects sheet: {e}")
        
        try:
            generate_certifications_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating certifications sheet: {e}")
        
        try:
            generate_achievements_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating achievements sheet: {e}")
        
        try:
            generate_placement_analytics_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating placement analytics sheet: {e}")
        
        try:
            generate_skills_analytics_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating skills analytics sheet: {e}")
        
        try:
            generate_naac_metrics_sheet(wb, students_queryset)
        except Exception as e:
            print(f"Error generating NAAC metrics sheet: {e}")
        
        # Ensure we have at least one sheet
        if len(wb.worksheets) == 0:
            ws = wb.create_sheet(title="Error")
            ws.cell(row=1, column=1, value="Error: No data could be generated")
        
        # Create a BytesIO object to save the workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="NAAC_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        import traceback
        error_msg = f"Error generating NAAC report: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)  # For debugging
        return JsonResponse({'error': error_msg}, status=500)

def generate_student_summary_sheet(wb, students_queryset):
    """Generate student summary sheet for NAAC report"""
    ws = wb.create_sheet(title="Student Summary")
    
    # Headers
    headers = [
        'Roll Number', 'Name', 'Department', 'Year', 'Section', 'Email',
        'CGPA', 'Projects Count', 'Certificates Count', 'Achievements Count',
        'LeetCode Problems', 'Mentor', 'Phone', 'GitHub Link'
    ]
    
    # Apply header styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        try:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        except Exception as e:
            # If styling fails, just set the value
            pass
    
    # Data rows
    for row, student_obj in enumerate(students_queryset, 2):
        try:
            # Get counts with error handling
            projects_count = 0
            certificates_count = 0
            achievements_count = 0
            
            try:
                projects_count = Projects.objects.filter(contributors=student_obj).count()
            except:
                pass
            
            try:
                certificates_count = Certificate.objects.filter(rollno=student_obj).count()
            except:
                pass
            
            try:
                achievements_count = Achievement.objects.filter(student=student_obj).count()
            except:
                pass
            
            # Get LeetCode count
            leetcode_count = 0
            try:
                if hasattr(student_obj, 'studentrollno') and student_obj.studentrollno:
                    leetcode_count = student_obj.studentrollno.TotalProblems or 0
            except:
                pass
            
            # Get mentor info safely
            mentor_name = 'N/A'
            try:
                if student_obj.mentor:
                    mentor_name = f"{student_obj.mentor.first_name} {student_obj.mentor.last_name}".strip()
            except:
                pass
            
            data = [
                getattr(student_obj, 'roll_no', 'N/A'),
                f"{getattr(student_obj, 'first_name', '')} {getattr(student_obj, 'last_name', '')}".strip() or 'N/A',
                getattr(student_obj, 'dept', 'N/A'),
                getattr(student_obj, 'year', 'N/A'),
                getattr(student_obj, 'section', 'N/A'),
                getattr(student_obj, 'email', 'N/A'),
                getattr(student_obj, 'current_cgpa', 'N/A') or 'N/A',
                projects_count,
                certificates_count,
                achievements_count,
                leetcode_count,
                mentor_name,
                getattr(student_obj, 'phone', 'N/A') or 'N/A',
                getattr(student_obj, 'github_link', 'N/A') or 'N/A'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                try:
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if col > 7:  # Numeric columns
                        cell.alignment = Alignment(horizontal="center")
                except:
                    pass
        except Exception as e:
            # If there's an error with this student, add an error row
            ws.cell(row=row, column=1, value=f"Error processing student: {str(e)}")
    
    # Auto-adjust column widths
    try:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        # If auto-adjustment fails, continue without it
        pass

def generate_academic_performance_sheet(wb, students_queryset):
    """Generate academic performance analytics sheet"""
    ws = wb.create_sheet(title="Academic Performance")
    
    # Performance summary data
    dept_stats = {}
    year_stats = {}
    
    for student_obj in students_queryset:
        dept = student_obj.dept
        year = student_obj.year
        cgpa = student_obj.current_cgpa or 0
        
        # Department statistics
        if dept not in dept_stats:
            dept_stats[dept] = {'count': 0, 'total_cgpa': 0, 'students': []}
        dept_stats[dept]['count'] += 1
        dept_stats[dept]['total_cgpa'] += float(cgpa) if cgpa != 'N/A' and cgpa else 0
        dept_stats[dept]['students'].append(student_obj)
        
        # Year statistics
        if year not in year_stats:
            year_stats[year] = {'count': 0, 'total_cgpa': 0}
        year_stats[year]['count'] += 1
        year_stats[year]['total_cgpa'] += float(cgpa) if cgpa != 'N/A' and cgpa else 0
    
    # Department-wise performance
    ws.cell(row=1, column=1, value="Department-wise Academic Performance").font = Font(size=14, bold=True)
    headers = ['Department', 'Student Count', 'Average CGPA', 'Projects/Student', 'Certificates/Student']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    row = 4
    for dept, stats in dept_stats.items():
        avg_cgpa = stats['total_cgpa'] / stats['count'] if stats['count'] > 0 else 0
        
        # Calculate projects and certificates per student
        total_projects = sum(Projects.objects.filter(contributors=s).count() for s in stats['students'])
        total_certificates = sum(Certificate.objects.filter(rollno=s).count() for s in stats['students'])
        
        projects_per_student = total_projects / stats['count'] if stats['count'] > 0 else 0
        certificates_per_student = total_certificates / stats['count'] if stats['count'] > 0 else 0
        
        ws.cell(row=row, column=1, value=dept)
        ws.cell(row=row, column=2, value=stats['count'])
        ws.cell(row=row, column=3, value=round(avg_cgpa, 2))
        ws.cell(row=row, column=4, value=round(projects_per_student, 1))
        ws.cell(row=row, column=5, value=round(certificates_per_student, 1))
        row += 1

def generate_projects_sheet(wb, students_queryset):
    """Generate projects analysis sheet"""
    ws = wb.create_sheet(title="Projects Analysis")
    
    # Get all projects for selected students
    all_projects = Projects.objects.filter(contributors__in=students_queryset).distinct()
    
    headers = [
        'Project Title', 'Description', 'Status', 'Year/Semester', 'Contributors Count',
        'Technologies', 'GitHub Link', 'Department'
    ]
    
    # Apply header styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Data rows
    for row, project in enumerate(all_projects, 2):
        contributors = project.contributors.all()
        contributor_names = ', '.join([f"{c.first_name} {c.last_name}" for c in contributors])
        technologies = ', '.join([tech.name for tech in project.technologies.all()])
        departments = ', '.join(list(set([c.dept for c in contributors])))
        
        data = [
            project.title,
            project.description[:100] + '...' if len(project.description) > 100 else project.description,
            project.status,
            project.year_and_sem,
            contributors.count(),
            technologies,
            project.github_link or 'N/A',
            departments
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

def generate_certifications_sheet(wb, students_queryset):
    """Generate certifications analysis sheet"""
    ws = wb.create_sheet(title="Certifications Analysis")
    
    try:
        # Get all certificates for selected students
        all_certificates = Certificate.objects.filter(rollno__in=students_queryset)
    except Exception as e:
        ws.cell(row=1, column=1, value=f"Error accessing certificate data: {str(e)}")
        return
    
    headers = [
        'Student Roll No', 'Student Name', 'Certificate Title', 'Category',
        'Source/Provider', 'Domain', 'Year/Semester', 'Validity Period'
    ]
    
    # Apply header styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        try:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        except:
            pass
    
    # Data rows
    try:
        for row, cert in enumerate(all_certificates, 2):
            try:
                data = [
                    getattr(cert.rollno, 'roll_no', 'N/A') if cert.rollno else 'N/A',
                    f"{getattr(cert.rollno, 'first_name', '')} {getattr(cert.rollno, 'last_name', '')}".strip() if cert.rollno else 'N/A',
                    getattr(cert, 'title', 'N/A'),
                    cert.get_category_display() if hasattr(cert, 'get_category_display') else getattr(cert, 'category', 'N/A'),
                    getattr(cert, 'source', 'N/A') or 'N/A',
                    getattr(cert, 'domain', 'N/A') or 'N/A',
                    getattr(cert, 'year_and_sem', 'N/A'),
                    getattr(cert, 'validity_period', 'N/A') or 'N/A'
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            except Exception as e:
                ws.cell(row=row, column=1, value=f"Error processing certificate: {str(e)}")
    except Exception as e:
        ws.cell(row=2, column=1, value=f"Error processing certificates: {str(e)}")
    
    # If no certificates found
    if all_certificates.count() == 0:
        ws.cell(row=2, column=1, value="No certificates found for selected students")

def generate_achievements_sheet(wb, students_queryset):
    """Generate achievements analysis sheet"""
    ws = wb.create_sheet(title="Achievements Analysis")
    
    # Get all achievements for selected students
    all_achievements = Achievement.objects.filter(student__in=students_queryset)
    
    headers = [
        'Student Roll No', 'Student Name', 'Achievement Title', 'Category',
        'Description', 'Achievement Date', 'Status', 'Points Awarded',
        'Verification Method'
    ]
    
    # Apply header styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    
    # Data rows
    for row, achievement in enumerate(all_achievements, 2):
        data = [
            achievement.student.roll_no,
            f"{achievement.student.first_name} {achievement.student.last_name}",
            achievement.title,
            achievement.category.name if achievement.category else 'N/A',
            achievement.description[:100] + '...' if len(achievement.description) > 100 else achievement.description,
            achievement.achievement_date.strftime('%Y-%m-%d') if achievement.achievement_date else 'N/A',
            achievement.get_status_display() if hasattr(achievement, 'get_status_display') else achievement.status,
            achievement.points_awarded or 0,
            achievement.verification_method or 'N/A'
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

def generate_placement_analytics_sheet(wb, students_queryset):
    """Generate placement analytics sheet"""
    ws = wb.create_sheet(title="Placement Analytics")
    
    try:
        # Get placement data
        placement_offers = PlacementOffer.objects.filter(student__in=students_queryset)
    except Exception as e:
        # If there's an error with placement data, create empty sheet with message
        ws.cell(row=1, column=1, value="No placement data available or error accessing placement records")
        return
    
    headers = [
        'Student Roll No', 'Student Name', 'Company', 'Package (LPA)',
        'Position', 'Placement Date', 'Department', 'Year'
    ]
    
    # Apply header styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    
    # Data rows
    try:
        for row, offer in enumerate(placement_offers, 2):
            data = [
                offer.student.roll_no if offer.student else 'N/A',
                f"{offer.student.first_name} {offer.student.last_name}" if offer.student else 'N/A',
                getattr(offer, 'company_name', 'N/A'),
                getattr(offer, 'package_lpa', 'N/A'),
                getattr(offer, 'position', 'N/A'),
                offer.placement_date.strftime('%Y-%m-%d') if hasattr(offer, 'placement_date') and offer.placement_date else 'N/A',
                offer.student.dept if offer.student else 'N/A',
                offer.student.year if offer.student else 'N/A'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
    except Exception as e:
        # If data access fails, add error message
        ws.cell(row=2, column=1, value=f"Error accessing placement data: {str(e)}")
    
    # If no data, add message
    if placement_offers.count() == 0:
        ws.cell(row=2, column=1, value="No placement offers found for selected students")

def generate_skills_analytics_sheet(wb, students_queryset):
    """Generate skills and technology analytics sheet"""
    ws = wb.create_sheet(title="Skills Analytics")
    
    # Technology usage statistics
    tech_stats = {}
    domain_stats = {}
    
    for student_obj in students_queryset:
        # Get technologies from projects
        student_projects = Projects.objects.filter(contributors=student_obj)
        for project in student_projects:
            for tech in project.technologies.all():
                tech_stats[tech.name] = tech_stats.get(tech.name, 0) + 1
        
        # Get domains from certificates
        student_certificates = Certificate.objects.filter(rollno=student_obj)
        for cert in student_certificates:
            if cert.domain:
                domain_stats[cert.domain] = domain_stats.get(cert.domain, 0) + 1
    
    # Technology statistics
    ws.cell(row=1, column=1, value="Technology Usage Statistics").font = Font(size=14, bold=True)
    ws.cell(row=3, column=1, value="Technology").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Usage Count").font = Font(bold=True)
    
    row = 4
    for tech, count in sorted(tech_stats.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=tech)
        ws.cell(row=row, column=2, value=count)
        row += 1
    
    # Domain statistics
    start_row = row + 2
    ws.cell(row=start_row, column=1, value="Certificate Domain Statistics").font = Font(size=14, bold=True)
    ws.cell(row=start_row + 2, column=1, value="Domain").font = Font(bold=True)
    ws.cell(row=start_row + 2, column=2, value="Certificate Count").font = Font(bold=True)
    
    row = start_row + 3
    for domain, count in sorted(domain_stats.items(), key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=domain)
        ws.cell(row=row, column=2, value=count)
        row += 1

def generate_naac_metrics_sheet(wb, students_queryset):
    """Generate NAAC-specific metrics sheet"""
    ws = wb.create_sheet(title="NAAC Metrics")
    
    total_students = students_queryset.count()
    
    # Calculate key metrics with proper error handling
    try:
        students_with_projects = 0
        students_with_certificates = 0
        students_with_achievements = 0
        students_with_placements = 0
        
        for student_obj in students_queryset:
            # Check projects
            if Projects.objects.filter(contributors=student_obj).exists():
                students_with_projects += 1
            
            # Check certificates
            if Certificate.objects.filter(rollno=student_obj).exists():
                students_with_certificates += 1
            
            # Check achievements
            if Achievement.objects.filter(student=student_obj).exists():
                students_with_achievements += 1
            
            # Check placements
            if PlacementOffer.objects.filter(student=student_obj).exists():
                students_with_placements += 1
        
        total_projects = Projects.objects.filter(contributors__in=students_queryset).distinct().count()
        total_certificates = Certificate.objects.filter(rollno__in=students_queryset).count()
        total_achievements = Achievement.objects.filter(student__in=students_queryset).count()
        
    except Exception as e:
        # Fallback values if queries fail
        students_with_projects = 0
        students_with_certificates = 0
        students_with_achievements = 0
        students_with_placements = 0
        total_projects = 0
        total_certificates = 0
        total_achievements = 0
    
    # Average calculations
    avg_projects_per_student = total_projects / total_students if total_students > 0 else 0
    avg_certificates_per_student = total_certificates / total_students if total_students > 0 else 0
    avg_achievements_per_student = total_achievements / total_students if total_students > 0 else 0
    
    # Create metrics table
    metrics = [
        ["NAAC Key Performance Indicators", ""],
        ["", ""],
        ["Total Students", total_students],
        ["Students with Projects", f"{students_with_projects} ({(students_with_projects/total_students*100):.1f}%)" if total_students > 0 else "0"],
        ["Students with Certificates", f"{students_with_certificates} ({(students_with_certificates/total_students*100):.1f}%)" if total_students > 0 else "0"],
        ["Students with Achievements", f"{students_with_achievements} ({(students_with_achievements/total_students*100):.1f}%)" if total_students > 0 else "0"],
        ["Students with Placements", f"{students_with_placements} ({(students_with_placements/total_students*100):.1f}%)" if total_students > 0 else "0"],
        ["", ""],
        ["Total Projects", total_projects],
        ["Total Certificates", total_certificates],
        ["Total Achievements", total_achievements],
        ["", ""],
        ["Average Projects per Student", f"{avg_projects_per_student:.2f}"],
        ["Average Certificates per Student", f"{avg_certificates_per_student:.2f}"],
        ["Average Achievements per Student", f"{avg_achievements_per_student:.2f}"],
    ]
    
    for row, (metric, value) in enumerate(metrics, 1):
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        
        if row == 1:  # Title row
            ws.cell(row=row, column=1).font = Font(size=16, bold=True)
        elif metric and value != "":  # Data rows
            ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

